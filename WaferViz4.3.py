# Started learning python 10/23/2021
# WaferViz4.1: 11/24/2023
# WaferViz4.2: 3/20/2024

import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.filedialog import askopenfilename
from PIL import ImageTk, Image
import tempfile, base64, zlib

from io import BytesIO
import win32clipboard
import win32api
import csv
import openpyxl

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
pd.options.mode.chained_assignment = None  

from matplotlib import gridspec
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from mpl_toolkits.mplot3d import axes3d
from matplotlib import cm
from scipy import stats

import math
from numpy import arccos, array
import os
import  sys
import threading

from scipy.interpolate import griddata
from matplotlib.patches import Circle
from matplotlib.figure import Figure

plt.rcParams['font.family'] ='Arial'
plt.rcParams['font.size'] = 5
plt.rcParams['axes.linewidth'] = 0.2
plt.rcParams['figure.dpi'] = 250
        
class app_gui:
    def __init__(self):
        ICON = zlib.decompress(base64.b64decode('eJxjYGAEQgEBBiDJwZDBy'
                                        'sAgxsDAoAHEQCEGBQaIOAg4sDIgACMUj4JRMApGwQgF/ykEAFXxQRc='))
        _, ICON_PATH = tempfile.mkstemp()
        with open(ICON_PATH, 'wb') as icon_file:
            icon_file.write(ICON)

        self.window_main = tk.Tk()
        self.window_main.iconbitmap(default=ICON_PATH)
        
        self.window_main.title("WaferViz4.3 - A Free Wafer Plotter - https://houli1959.github.io/plasma/")
        self.window_main.geometry("1596x870+150+70")
        self.window_main.resizable(1, 1)
        self.window_main.minsize(300, 300)
        self.window_main.columnconfigure(0, weight=1)
        self.window_main.rowconfigure(1, weight=1)

        self.frame_control = tk.Frame(self.window_main, borderwidth=1, relief=SUNKEN)
        self.frame_control.grid(row=0, column=0, sticky="nesw", padx=1, pady=0)
        self.frame_control.columnconfigure(0, weight=0)
        self.frame_control.rowconfigure(0, weight=0)

        self.frame_display = tk.Frame(self.window_main, borderwidth=1, relief=SUNKEN)
        self.frame_display.grid(row=1, column=0, sticky="nesw", padx=1, pady=0)
        self.frame_display.columnconfigure(0, weight=1)
        self.frame_display.rowconfigure(0, weight=1)

        self.button_workbook = tk.Button(self.frame_control, text="Load Data Sheet",
                            command=self.open_workbook, width=16)     
        self.button_workbook.grid(row=0, column=0, sticky="e", columnspan=1, padx=10, pady=2)

        self.button_atlas = tk.Button(self.frame_control, text="Load Atlas Exported",
                            command=self.open_atlas, width=16)     
        self.button_atlas.grid(row=1, column=0, sticky="e", columnspan=1, padx=10, pady=2)

        self.button_atlas_map = tk.Button(self.frame_control, text="Load Atlas Map",
                            command=self.open_atlas_map, width=16)     
        self.button_atlas_map.grid(row=2, column=0, sticky="e", columnspan=1, padx=10, pady=2)

        self.button_resmap = tk.Button(self.frame_control, text="Load CDE ResMap",
                            command=self.open_resmap, width=16)     
        self.button_resmap.grid(row=3, column=0, sticky="e", columnspan=1, padx=10, pady=2)

        option_sel = ["Thickness", "MSE", "Roughness"]
        self.variable_sel = tk.StringVar()
        self.variable_sel.set("Thickness")

        self.label_id = tk.Label(self.frame_control, text="Enter Id")
        self.label_id.grid(row=2, column=1, sticky='e', padx=2, pady=0)
        self.entry_id = tk.Entry(self.frame_control, width=12, justify='left')
        self.entry_id.insert(0, ' ')
        self.entry_id.grid(row=2, column=2, columnspan=1, sticky='w', padx=0, pady=0)

        self.file_path = tk.Label(self.frame_control, text="", justify='left',  anchor="w", fg="#717171",
                           highlightthickness=0, font=("Arial", 7), wraplength=460)
        self.file_path.grid(row=3, column=1, columnspan=7, sticky="w", padx=0, pady=0)

        self.selected_a = tk.StringVar()
        self.so1 = ttk.Combobox(self.frame_control, textvariable=self.selected_a, justify='center')
        self.so1.config(width=10)
        self.so1['state'] = 'readonly'
        self.so1.grid(column=5, row=0,  sticky='w', padx=0, pady=0)
        self.label_combo1 = tk.Label(self.frame_control, text="A")
        self.label_combo1.grid(row=0, column=4, sticky='e', padx=2, pady=0)

        self.selected_b = tk.StringVar()
        self.so2 = ttk.Combobox(self.frame_control, textvariable=self.selected_b, justify='center')
        self.so2.config(width=10)
        self.so2['state'] = 'readonly'
        self.so2.grid(column=7, row=0,  sticky='w', padx=0, pady=0)
        self.label_combo2 = tk.Label(self.frame_control, text="B")
        self.label_combo2.grid(row=0, column=6, sticky='e', padx=2, pady=0)

        option_calist = ["A-B", "B-A", "(A-B)/t", "(B-A)/t", "A", "B", "A/t", "B/t"]
        self.variable_cal = tk.StringVar()
        self.variable_cal.set("A-B")
        self.option_calbtn = tk.OptionMenu(self.frame_control, self.variable_cal, *option_calist)
        self.option_calbtn.config(width=5)
        self.option_calbtn.grid(column=5, row=1, columnspan=2, sticky='w', padx=0, pady=0)

        self.entry_run_time = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_run_time.insert(0, 60)
        self.entry_run_time.grid(row=3, column=10, sticky='w', padx=2, pady=0)
        self.label_run_time = tk.Label(self.frame_control, text="Set Time (s)")
        self.label_run_time.grid(row=3, column=9, sticky='e', padx=0, pady=0)

        self.label_rotation = tk.Label(self.frame_control, width=10, text="Set Rotation", anchor="e")
        self.label_rotation.grid(row=0, column=9, sticky='e', padx=0, pady=0)
        self.rotation_entry = tk.Entry(self.frame_control, width=6, justify='center')
        self.rotation_entry.insert(0, 0)
        self.rotation_entry.grid(row=0, column=10, sticky='w', padx=4, pady=0)

        self.label_contour = tk.Label(self.frame_control, width=10, text="Set Contours", anchor="e")
        self.label_contour.grid(row=1, column=9, sticky='e', padx=2, pady=0)
        self.entry_contour = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_contour.insert(0, 10)
        self.entry_contour.grid(row=1, column=10, sticky='w', padx=2, pady=0)

        self.entry_unit = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_unit.insert(0, 'Å')
        self.entry_unit.grid(row=2, column=10, sticky='w', padx=2, pady=0)
        self.label_unit = tk.Label(self.frame_control, text="Set Unit", anchor="e")
        self.label_unit.grid(row=2, column=9, sticky='e', padx=2, pady=0)

        self.radio_sigma = tk.IntVar()
        self.radio_sigma.set(2)
        self.button_sigma = tk.Radiobutton(self.frame_control, text='Control Limits \u00B1\u03C3',
                            variable=self.radio_sigma, value=1, indicator=0, width=14)
        self.button_sigma.grid(column=13, row=1, columnspan=1, sticky='e', padx=2, pady=0)
        self.entry_sigma = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_sigma.insert(0, 3)
        self.entry_sigma.grid(row=1, column=14, sticky='w', padx=4, pady=0)
        self.button_outlier = tk.Radiobutton(self.frame_control, text='Remove Outliers', 
                            variable=self.radio_sigma, value=2, indicator=0, width=14)  
        self.button_outlier.grid(column=13, row=0, columnspan=1, sticky='e', padx=2, pady=0)
        self.entry_outlier = tk.Entry(self.frame_control, width=6, justify='center')
        self.entry_outlier.insert(0, 0)
        self.entry_outlier.grid(row=0, column=14, sticky='w', padx=4, pady=0)

        option_list3 = ["Sign", "Value", "Dot", "None"]
        self.variable3 = tk.StringVar()
        self.variable3.set("Sign")
        self.option_button3 = tk.OptionMenu(self.frame_control, self.variable3, *option_list3)
        self.option_button3.config(width=5)
        self.option_button3.grid(column=14, row=3,  sticky='w', padx=1, pady=0)
        self.marker = tk.Label(self.frame_control, width=10, text="Map Markers", anchor="e")
        self.marker.grid(row=3, column=13, sticky='e', padx=2, pady=0)

        self.spacer1 = tk.Label(self.frame_control, width=5, text="")      
        self.spacer1.grid(row=0, column=3, sticky='ew')
        
        self.spacer1 = tk.Label(self.frame_control, width=5, text="")      
        self.spacer1.grid(row=0, column=8, sticky='ew')
        
        self.spacer2 = tk.Label(self.frame_control, width=1, text="")        
        self.spacer2.grid(row=0, column=20, sticky='ew')
        
        self.spacer3 = tk.Label(self.frame_control, width=8, text="")        
        self.spacer3.grid(row=1, column=11, sticky='ew')
        
        self.spacer4 = tk.Label(self.frame_control, width=8, text="")        
        self.spacer4.grid(row=1, column=15, sticky='ew')
        
        self.spacer5 = tk.Label(self.frame_control, width=3, text="")        
        self.spacer5.grid(row=1, column=17, sticky='ew')
        
        self.spacer6 = tk.Label(self.frame_control, width=5, text="")        
        self.spacer6.grid(row=1, column=21, sticky='ew')

        self.var_limits = tk.IntVar()
        self.var_limits.set(2)
        self.button_range = tk.Radiobutton(self.frame_control, text='Auto Limits', variable=self.var_limits,
                            value=2, indicator=0, width=11)  
        self.button_range.grid(column=16, row=0, columnspan=1, sticky='e', padx=4, pady=0)
        self.button_ul = tk.Radiobutton(self.frame_control, text='Set Limits', variable=self.var_limits, 
                            value=1, indicator=0, width=11)
        self.button_ul.grid(column=16, row=1, columnspan=1, sticky='e', padx=4, pady=0)
        
        self.limits_entry = tk.Entry(self.frame_control, width=6, justify='center')
        self.limits_entry.insert(0, '200')
        self.limits_entry.grid(row=1, column=17, sticky='e', padx=4, pady=0)
        self.limits_up = tk.Entry(self.frame_control, width=6, justify='center')
        self.limits_up.insert(0, '2000')
        self.limits_up.grid(row=1, column=18, sticky='w', padx=2, pady=0)

        self.var_decom = tk.IntVar()
        self.check_decom = tk.Checkbutton(self.frame_control, text='Decompose',
                            variable=self.var_decom)
        self.check_decom.grid(row=3, column=16, sticky='w', padx=0, pady=0, columnspan=2)
        
        self.var_animated = tk.IntVar()
        self.check_animated = tk.Checkbutton(self.frame_control, text='3d Animated',
                            variable=self.var_animated)
        self.check_animated.grid(row=3, column=17, sticky='e', padx=10, pady=0, columnspan=2)
        
        self.button_ps = tk.Button(self.frame_control, text="Plot", command=self.save_file, width=13)
        self.button_ps.grid(row=0, column=23, sticky="e", padx=4, pady=0)
        
        self.button_copy = tk.Button(self.frame_control, text="How to Use",
                            command=self.open_pdf, width=13)
        self.button_copy.grid(row=3, column=23, sticky="e", padx=4, pady=0)
        
        self.canvas_graph = tk.Canvas(self.frame_display, bg='white', bd=0, scrollregion=[0,0,1650,1600])
        self.canvas_graph.grid(row=0, column=0, sticky="nesw")
        self.canvas_graph.columnconfigure(0, weight=1)
        self.canvas_graph.rowconfigure(0, weight=1)

        self.scrollbar_vertical = tk.Scrollbar(self.frame_display, orient="vertical", relief=tk.SUNKEN, bd=0,
                            width=23)
        self.scrollbar_vertical.grid(row=0, column=1, sticky="ns")
        self.canvas_graph.configure(yscrollcommand=self.scrollbar_vertical.set)
        self.scrollbar_vertical.config(command=self.canvas_graph.yview)

        self.scrollbar_horizontal = tk.Scrollbar(self.frame_display, orient="horizontal", relief=tk.SUNKEN,
                            bd=0, width=23)
        self.scrollbar_horizontal.grid(row=1, column=0, sticky="ew")
        self.canvas_graph.configure(xscrollcommand=self.scrollbar_horizontal.set)
        self.scrollbar_horizontal.config(command=self.canvas_graph.xview)
        
        tk.mainloop()

    def open_pdf(self):
        import fitz
        input_file = r"C:\Users\HouL\AppData\Local\Programs\Python\Python311\Scripts\How to use.pdf"
        file_handle = fitz.open(input_file)
        page = file_handle[0]
        page_img = page.get_pixmap(dpi=450)
        page_img.save('PDF_page.png')
        img = Image.open('PDF_page.png')
        img.show()

        tk.mainloop()
        
    def open_atlas(self):
        global df1
        global df3
        global df32
        global df3_real
        global index_da
        global index_a
        global index_b
        index_da = 1
        index_a = 0
        index_b = 0
        
        filepath = askopenfilename(filetypes=[("All", "*.*"), ("CSV", "*.csv")])
        path = filepath
        self.file_path.configure(text=path)
        filename = os.path.basename(filepath)

        extension = os.path.splitext(filename)[1]
        if (extension == '.csv'):
            def csv_to_excel(csv_file, excel_file):
                csv_data = []
                with open(csv_file) as file_obj:
                    reader = csv.reader(file_obj)
                    for row in reader:
                        csv_data.append(row)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                for row in csv_data:
                    sheet.append(row)
                workbook.save(excel_file)
            if __name__ == "__main__":
                csv_to_excel(filepath, './tmp.xlsx')
                df_data = pd.read_excel('./tmp.xlsx')
                os.remove('tmp.xlsx')
        elif (extension == '.xlsx'):
            df_data = pd.read_excel(filepath)
        else:
            win32api.MessageBox(0, 'You can only select files with either .csv or .xlsx', 'Error:')

        df1 = pd.DataFrame(df_data) 
        cols = len(df1.axes[1]) + 1
        column_names = pd.Series(np.arange(1, cols, 1))
        
        mask = np.column_stack([df1[col].str.contains(r"Wafer ID:", na=False) for col in df1])
        df2 = df1.loc[mask.any(axis=1)]
        df2.columns = column_names
        df3 = (df2.loc[:, 2]).reset_index()
        col2_list = df3[2].values.tolist()
        
        mask_real = np.column_stack([df1[col].str.contains(r"Title:", na=False) for col in df1])
        df2_real = df1.loc[mask_real.any(axis=1)]
        df2_real.columns = column_names
        df3_real = (df2_real.loc[:, 2]).reset_index()
        col2_list_real = df3_real['index'].values.tolist()
        
        mask2 = np.column_stack([df1[col].str.contains(r"Max", na=False) for col in df1])
        df22 = df1.loc[mask2.any(axis=1)]
        df22.columns = column_names
        df32 = (df22.loc[:, 2]).reset_index()

        variable1 = tk.StringVar()
        self.cb1 = ttk.Combobox(self.frame_control, textvariable=variable1, justify='center', height=40)
        self.cb1.bind('<<ComboboxSelected>>', self.callback1)
        self.cb1.grid(column=5, row=0,  sticky='w', padx=0, pady=0)
        self.cb1.config(values=col2_list, width=10)
        self.cb1.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb1['state']= 'readonly'

        variable2 = tk.StringVar()
        self.cb2 = ttk.Combobox(self.frame_control, textvariable=variable2, justify='center', height=40)
        self.cb2.bind('<<ComboboxSelected>>', self.callback2)
        self.cb2.grid(column=7, row=0,  sticky='w', padx=0, pady=0)
        self.cb2.config(values=col2_list, width=10)
        self.cb2.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb2['state']= 'readonly'
        
    def callback1(self, *args):
        global position_a
        global index_a
        index_a = 1
        position_a = str(self.cb1.current())

    def callback2(self, *args):
        global position_b
        global index_b
        index_b = 2
        position_b = str(self.cb2.current())
        
        tk.mainloop()

    def open_workbook(self):
        global df2
        global row_end
        global index_da
        global index_a
        global index_b
        index_da = 0
        index_a = 0
        index_b = 0

        filepath = askopenfilename(filetypes=[("All", "*.*"), ("Excel", "*.xlsx")])
        path = filepath
        self.file_path.configure(text=path)
        filename = os.path.basename(filepath)

        extension = os.path.splitext(filename)[1]
        if (extension == '.csv'):
            def csv_to_excel(csv_file, excel_file):
                csv_data = []
                with open(csv_file) as file_obj:
                    reader = csv.reader(file_obj)
                    for row in reader:
                        csv_data.append(row)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                for row in csv_data:
                    sheet.append(row)
                workbook.save(excel_file)
            if __name__ == "__main__":
                csv_to_excel(filepath, './tmp.xlsx')
                df_data = pd.read_excel('./tmp.xlsx')
                os.remove('tmp.xlsx')
        elif (extension == '.xlsx'):
            df_data = pd.read_excel(filepath)
        else:
            win32api.MessageBox(0, 'You can only select files with either .csv or .xlsx', 'Error:')

        df2 = pd.DataFrame(df_data)
        nan_value = float("NaN")
        df2.replace("", nan_value, inplace=True)
        df2.dropna(how='all', axis=1, inplace=True)
        col_name = list(df2.columns.values)
        col_name1 = [s.replace('Unnamed:', '') for s in col_name]

        def make_int(s):
            if not s:
                return s
            try:
                f = float(s)
                i = int(f)
                return i if f == i else f
            except ValueError:
                return s
        converted = list(map(make_int, col_name1))
        converted = [t + 1 if type(t) in [int, float, bool] else t for t in converted]
        del converted[0:2]

        output_list = []
        for element in converted:
            value = str(element)
            output_list.append(value)

        prefix = '# '
        pre_res = []
        for item in output_list:
            pre_res.append(prefix + item)

        variable3 = tk.StringVar()
        self.cb3 = ttk.Combobox(self.frame_control, textvariable=variable3, justify='center', height=40)
        self.cb3.bind('<<ComboboxSelected>>', self.callback3)
        self.cb3.grid(column=5, row=0,  sticky='w', padx=0, pady=0)
        self.cb3.config(values=pre_res, width=10)
        self.cb3.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb3['state']= 'readonly'

        variable4 = tk.StringVar()
        self.cb4 = ttk.Combobox(self.frame_control, textvariable=variable4, justify='center', height=40)
        self.cb4.bind('<<ComboboxSelected>>', self.callback4)
        self.cb4.grid(column=7, row=0,  sticky='w', padx=0, pady=0)
        self.cb4.config(values=pre_res, width=10)
        self.cb4.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb4['state']= 'readonly'

        row_num = len(df2.index)
        row_end = row_num +1
        
    def callback3(self, *args):
        global dfA_cb34
        global index_a
        index_a = 1
        column_a = int(str(self.cb3.current())) +2
        dfA_cb34 = df2.iloc[0:row_end, column_a]

    def callback4(self, *args):
        global dfB_cb34
        global index_b
        index_b = 2
        column_b = int(str(self.cb4.current())) +2
        dfB_cb34 = df2.iloc[0:row_end, column_b]
        
        tk.mainloop()

    def open_atlas_map(self):
        global df_pre
        global df2
        global row_end
        global index_da
        global index_a
        global index_b
        
        index_da = 2
        index_a = 0
        index_b = 0
        filepath = askopenfilename(filetypes=[("All", "*.*"), ("Map", "*.map")])
        path = filepath
        self.file_path.configure(text=path)
        filename = os.path.basename(filepath)
        
        extension = os.path.splitext(filename)[1]
        if (extension == '.map'):
            f = open(filepath)
            df_pre = pd.read_csv(f, skiprows=0, sep=' ', header=None, names=["a1", "a2", "a3", "a4",
                            "a5", "a6", "a7", "a8", "a9", "a10", "a11", "a12", "a13", "a14", "a15", "a16", "a17",
                            "a18", "a19", "a20", "a21", "a22", "a23", "a24", "a25"])  
            c = df_pre.index[(df_pre['a1']=='Data') & (df_pre['a2']=='Section:')].tolist()
            def convert(list):
                s = [str(i) for i in list]
                res = int("".join(s))
                return(res)
            list = c
            c1 = convert(list) + 2
            f = open(filepath)
            df_data = pd.read_csv(f, skiprows=c1, sep=' ', header=None, names=["zero", "X", "Y", "Z"])
            df_data = df_data.drop('zero', axis=1)
            
        elif (extension == '.txt'):
            f = open(filepath)
            df_pre = pd.read_csv(f, skiprows=0, sep=' ', header=None, names=["a1", "a2", "a3", "a4",
                            "a5", "a6", "a7", "a8", "a9", "a10", "a11", "a12", "a13", "a14", "a15", "a16", "a17",
                            "a18", "a19", "a20", "a21", "a22", "a23", "a24", "a25"])  
            c = df_pre.index[(df_pre['a1']=='Data') & (df_pre['a2']=='Section:')].tolist()
            
            def convert(list):
                s = [str(i) for i in list]
                res = int("".join(s))
                return(res)
            list = c
            c1 = convert(list) + 2
            f = open(filepath)
            df_data = pd.read_csv(f, skiprows=c1, sep=' ', header=None, names=["zero", "X", "Y", "Z"])
            df_data = df_data.drop('zero', axis=1)
            
        else:
            win32api.MessageBox(0, 'You can only select files with either .map or .txt', 'Error:')

        df2 = pd.DataFrame(df_data)

        variable5 = tk.StringVar()
        self.cb5 = ttk.Combobox(self.frame_control, textvariable=variable5, justify='center', height=40)
        self.cb5.bind('<<ComboboxSelected>>', self.callback5)
        self.cb5.grid(column=5, row=0,  sticky='w', padx=0, pady=0)
        self.cb5.config(values="Map1", width=10)
        self.cb5.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb5['state']= 'readonly'

        variable6 = tk.StringVar()
        self.cb6 = ttk.Combobox(self.frame_control, textvariable=variable6, justify='center', height=40)
        self.cb6.bind('<<ComboboxSelected>>', self.callback6)
        self.cb6.grid(column=7, row=0,  sticky='w', padx=0, pady=0)
        self.cb6.config(values="Map1", width=10)
        self.cb6.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb6['state']= 'readonly'

        row_num = len(df2.index)
        row_end = row_num +1
        
    def callback5(self, *args):
        global dfA_cb56
        global index_a
        index_a = 1
        column_a = int(str(self.cb5.current())) +2
        dfA_cb56 = df2.iloc[0:row_end, column_a]

    def callback6(self, *args):
        global dfB_cb56
        global index_b
        index_b = 2
        column_b = int(str(self.cb6.current())) +2
        dfB_cb56 = df2.iloc[0:row_end, column_b]
        
        tk.mainloop()

    def open_resmap(self):
        global df_pre
        global df2
        global row_end
        global index_da
        global index_a
        global index_b
        
        index_da = 3
        index_a = 0
        index_b = 0
        filepath = askopenfilename(filetypes=[("All", "*.*"), ("CSV", "*.csv"), ("Excel", "*.xlsx")])
        path = filepath
        self.file_path.configure(text=path)
        filename = os.path.basename(filepath)

        extension = os.path.splitext(filename)[1]
        if (extension == '.csv' or extension == '.CSV'):
            def csv_to_excel(csv_file, excel_file):
                csv_data = []
                with open(csv_file) as file_obj:
                    reader = csv.reader(file_obj)
                    for row in reader:
                        csv_data.append(row)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                for row in csv_data:
                    sheet.append(row)
                workbook.save(excel_file)
            if __name__ == "__main__":
                csv_to_excel(filepath, './tmp.xlsx')
                df_pre = pd.read_excel('./tmp.xlsx')
                os.remove('tmp.xlsx')
        elif (extension == '.xlsx'):
            df_pre = pd.read_excel(filepath)
        else:
            win32api.MessageBox(0, 'You can only select files with either .csv or .xlsx', 'Error:')

        df_pre1 = df_pre.set_index(df_pre.columns[0])
        df_pre1.index = [x for x in range(1, len(df_pre1.values)+1)]
        df_pre2 = df_pre1[['Unnamed: 7', 'Unnamed: 8', '<Title>']]
        c = df_pre1.index[(df_pre1['Unnamed: 15']=='<R')].tolist()
        def convert(list):
            s = [str(i) for i in list]
            res = int("".join(s))
            return(res)
        list = c
        c1 = convert(list)
        df_pre3 = df_pre2.rename(columns={'Unnamed: 7': 'X', 'Unnamed: 8': 'Y', '<Title>': 'Z'})
        df_pre4 = df_pre3.drop(df_pre2.index[range(c1)])
        df_pre5 = df_pre4.reset_index()
        df2 = df_pre5.drop(['index'], axis=1)

        variable7 = tk.StringVar()
        self.cb7 = ttk.Combobox(self.frame_control, textvariable=variable7, justify='center', height=40)
        self.cb7.bind('<<ComboboxSelected>>', self.callback7)
        self.cb7.grid(column=5, row=0,  sticky='w', padx=0, pady=0)
        self.cb7.config(values="Map1", width=10)
        self.cb7.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb7['state']= 'readonly'

        variable8 = tk.StringVar()
        self.cb8 = ttk.Combobox(self.frame_control, textvariable=variable8, justify='center', height=40)
        self.cb8.bind('<<ComboboxSelected>>', self.callback8)
        self.cb8.grid(column=7, row=0,  sticky='w', padx=0, pady=0)
        self.cb8.config(values="Map1", width=10)
        self.cb8.option_add('*TCombobox*Listbox.Justify', 'center')   
        self.cb8['state']= 'readonly'

        row_num = len(df2.index)
        row_end = row_num +1
        
    def callback7(self, *args):
        global dfA_cb78
        global index_a
        index_a = 1
        column_a = int(str(self.cb7.current())) +2
        dfA_cb78 = df2.iloc[0:row_end, column_a]

    def callback8(self, *args):
        global dfB_cb78
        global index_b
        index_b = 2
        column_b = int(str(self.cb8.current())) +2
        dfB_cb78 = df2.iloc[0:row_end, column_b]
        
        tk.mainloop()

    def save_file(self):
        global id
        global df
        global unit
        global value1
        global time
        global N1
        global N
        global df5
        
# save atlas
        if (index_da == 1):
            index = index_a + index_b        
            self.canvas_graph.delete('all')       
            id = self.entry_id.get()
            ID = id
            unit = self.entry_unit.get()

#for both positions
            if (index_a == 1 and index_b == 2):
                df4a = df3_real.iloc[[position_a]]
                row_num1 = df4a['index'].loc[df4a.index[0]]
                start1 = row_num1 + 1
                df421 = df32.iloc[[position_a]]
                row_num12 = df421['index'].loc[df421.index[0]]
                end1 = row_num12 - 1

                df_col = df1.loc[[start1-1]]
                cols = len(df_col.axes[1])
                values = range(cols)
                for i in values:
                    df_col.columns.values[i] = i
                lst = df_col.iloc[0].tolist()

                var_sel = self.variable_sel.get()
                if (var_sel == "Thickness"):
                    if 'Si3N4 (Å)' in lst:
                        col_num1 = lst.index('Si3N4 (Å)')
                    else: 
                        if 'Poly (Å)' in lst:
                            col_num1 = lst.index('Poly (Å)')
                        else:
                            if 'SiO2 (Å)' in lst:
                                col_num1 = lst.index('SiO2 (Å)')
                            else:
                                if 'Resist (Å)' in lst:
                                    col_num1 = lst.index('Resist (Å)')
                                else:
                                    if 'AC (Å)' in lst:
                                        col_num1 = lst.index('AC (Å)')
                                    else:
                                        win32api.MessageBox(0,
                            'No match amound Si3N4, Poly, SiO2, Resist and AC', 'Error:')
                else:
                    pass
         
                dfA = df1.iloc[start1:end1, col_num1]
                dfA1 = dfA.reset_index()
                dfA1.columns = ['rows_a', 'A']
                dfA1 = dfA1['A']

                col_num1x = lst.index('X(mm)')
                col_num1y = lst.index('Y(mm)')
        
                dfX = df1.iloc[start1:end1, col_num1x]
                dfX1 = dfX.reset_index()
                dfX1.columns = ['rows_x', 'X']
                dfX1 = dfX1['X']

                dfY = df1.iloc[start1:end1, col_num1y]
                dfY1 = dfY.reset_index()
                dfY1.columns = ['rows_y', 'Y']
                dfY1 = dfY1['Y']

                df4b = df3_real.iloc[[position_b]]
                row_num2 = df4b['index'].loc[df4b.index[0]]
                start2 = row_num2 + 1
                df422 = df32.iloc[[position_b]]
                row_num22 = df422['index'].loc[df422.index[0]]
                end2 = row_num22 - 1

                df_col = df1.loc[[start2-1]]
                cols = len(df_col.axes[1])
                values = range(cols)
                for i in values:
                    df_col.columns.values[i] = i
                lst = df_col.iloc[0].tolist()

                var_sel = self.variable_sel.get()
                if (var_sel == "Thickness"):
                    if 'Si3N4 (Å)' in lst:
                        col_num2 = lst.index('Si3N4 (Å)')
                    else: 
                        if 'Poly (Å)' in lst:
                            col_num2 = lst.index('Poly (Å)')
                        else:
                            if 'SiO2 (Å)' in lst:
                                col_num2 = lst.index('SiO2 (Å)')
                            else:
                                if 'Resist (Å)' in lst:
                                    col_num2 = lst.index('Resist (Å)')
                                else:
                                    if 'AC (Å)' in lst:
                                        col_num2 = lst.index('AC (Å)')
                                    else:
                                        win32api.MessageBox(0,
                            'No match amound Si3N4, Poly, SiO2, Resist and AC', 'Error:')
                                    
                elif (var_sel == "MSE"):
                    if 'MSE' in lst:
                        col_num2 = lst.index('MSE')
                    else:
                        win32api.MessageBox(0, 'No match found for MSE', 'Error:')
                
                elif (var_sel == "Roughness"):
                    if 'Roughness (Å)' in lst:
                        col_num2 = lst.index('Roughness (Å)')
                    else:
                        win32api.MessageBox(0, 'No match found for Roughness', 'Error:')
                else:
                    pass
        
                dfB = df1.iloc[start2:end2, col_num2]
                dfB1 = dfB.reset_index()      
                dfB1.columns = ['rows_b', 'B']
                dfB1 = dfB1['B']

                col_num2x = lst.index('X(mm)')
                col_num2y = lst.index('Y(mm)')

                dfX = df1.iloc[start2:end2, col_num2x]
                dfX2 = dfX.reset_index()
                dfX2.columns = ['rows_x', 'X']
                dfX2 = dfX2['X']
        
                dfY = df1.iloc[start2:end2, col_num2y]
                dfY2 = dfY.reset_index()
                dfY2.columns = ['rows_y', 'Y']
                dfY2 = dfY2['Y']
            
#for position a
            elif (index_a == 1):
                df4a = df3_real.iloc[[position_a]]
                row_num1 = df4a['index'].loc[df4a.index[0]]
                start1 = row_num1 + 1
                df421 = df32.iloc[[position_a]]
                row_num12 = df421['index'].loc[df421.index[0]]
                end1 = row_num12 - 1

                df_col = df1.loc[[start1-1]]
                cols = len(df_col.axes[1])
                values = range(cols)
                for i in values:
                    df_col.columns.values[i] = i
                lst = df_col.iloc[0].tolist()

                var_sel = self.variable_sel.get()
                if (var_sel == "Thickness"):
                    if 'Si3N4 (Å)' in lst:
                        col_num1 = lst.index('Si3N4 (Å)')
                    else: 
                        if 'Poly (Å)' in lst:
                            col_num1 = lst.index('Poly (Å)')
                        else:
                            if 'SiO2 (Å)' in lst:
                                col_num1 = lst.index('SiO2 (Å)')
                            else:
                                if 'Resist (Å)' in lst:
                                    col_num1 = lst.index('Resist (Å)')
                                else:
                                    if 'AC (Å)' in lst:
                                        col_num1 = lst.index('AC (Å)')
                                    else:
                                        win32api.MessageBox(0,
                            'No match amound Si3N4, Poly, SiO2, Resist and AC', 'Error:')
                                    
                elif (var_sel == "MSE"):
                    if 'MSE' in lst:
                        col_num1 = lst.index('MSE')
                    else:
                        win32api.MessageBox(0, 'No match found for MSE', 'Error:')
                
                elif (var_sel == "Roughness"):
                    if 'Roughness (Å)' in lst:
                        col_num1 = lst.index('Roughness (Å)')
                    else:
                        win32api.MessageBox(0, 'No match found for Roughness', 'Error:')
                else:
                    pass
         
                dfA = df1.iloc[start1:end1, col_num1]
                dfA1 = dfA.reset_index()
                dfA1.columns = ['rows_a', 'A']
                dfA1 = dfA1['A']

                col_num1x = lst.index('X(mm)')
                col_num1y = lst.index('Y(mm)')
        
                dfX = df1.iloc[start1:end1, col_num1x]
                dfX1 = dfX.reset_index()
                dfX1.columns = ['rows_x', 'X']
                dfX1 = dfX1['X']

                dfY = df1.iloc[start1:end1, col_num1y]
                dfY1 = dfY.reset_index()
                dfY1.columns = ['rows_y', 'Y']
                dfY1 = dfY1['Y']
            
#for position b
            elif (index_b == 2):
                df4b = df3_real.iloc[[position_b]]
                row_num2 = df4b['index'].loc[df4b.index[0]]
                start2 = row_num2 + 1
                df422 = df32.iloc[[position_b]]
                row_num22 = df422['index'].loc[df422.index[0]]
                end2 = row_num22 - 1

                df_col = df1.loc[[start2-1]]
                cols = len(df_col.axes[1])
                values = range(cols)
                for i in values:
                    df_col.columns.values[i] = i
                lst = df_col.iloc[0].tolist()

                var_sel = self.variable_sel.get()
                if (var_sel == "Thickness"):
                    if 'Si3N4 (Å)' in lst:
                        col_num2 = lst.index('Si3N4 (Å)')
                    else: 
                        if 'Poly (Å)' in lst:
                            col_num2 = lst.index('Poly (Å)')
                        else:
                            if 'SiO2 (Å)' in lst:
                                col_num2 = lst.index('SiO2 (Å)')
                            else:
                                if 'Resist (Å)' in lst:
                                    col_num2 = lst.index('Resist (Å)')
                                else:
                                    if 'AC (Å)' in lst:
                                        col_num2 = lst.index('AC (Å)')
                                    else:
                                        win32api.MessageBox(0,
                            'No match amound Si3N4, Poly, SiO2, Resist and AC', 'Error:')
                                    
                elif (var_sel == "MSE"):
                    if 'MSE' in lst:
                        col_num2 = lst.index('MSE')
                    else:
                        win32api.MessageBox(0, 'No match found for MSE', 'Error:')
                
                elif (var_sel == "Roughness"):
                    if 'Roughness (Å)' in lst:
                        col_num2 = lst.index('Roughness (Å)')
                    else:
                        win32api.MessageBox(0, 'No match found for Roughness', 'Error:')
                else:
                    pass
        
                dfB = df1.iloc[start2:end2, col_num2]
                dfB1 = dfB.reset_index()      
                dfB1.columns = ['rows_b', 'B']
                dfB1 = dfB1['B']

                col_num2x = lst.index('X(mm)')
                col_num2y = lst.index('Y(mm)')

                dfX = df1.iloc[start2:end2, col_num2x]
                dfX2 = dfX.reset_index()
                dfX2.columns = ['rows_x', 'X']
                dfX2 = dfX2['X']
        
                dfY = df1.iloc[start2:end2, col_num2y]
                dfY2 = dfY.reset_index()
                dfY2.columns = ['rows_y', 'Y']
                dfY2 = dfY2['Y']

            else:
                pass

            value1 = self.variable_cal.get()
            if value1 == 'A-B' or value1 == '(A-B)/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select films for "A" and "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:
                    df5 = pd.concat([dfX1, dfY1, dfA1, dfB1], axis=1)
                    df5.columns = ['X', 'Y', 'A', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5['B'] = df5['B'].apply(float)
                    
                    if df5['A'].equals(df5['B']):
                        win32api.MessageBox(0, 'You have selected the same film for "A" and "B"', 'Error:')
                    else:
                        df5["Z"] = df5["A"] - df5["B"]
                        
            elif value1 == 'B-A' or value1 == '(B-A)/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select films for "A" and "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:            
                    df5 = pd.concat([dfX1, dfY1, dfA1, dfB1], axis=1)
                    df5.columns = ['X', 'Y', 'A', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5['B'] = df5['B'].apply(float)

                    if df5['A'].equals(df5['B']):
                        win32api.MessageBox(0, 'You have selected the same film for "A" and "B"', 'Error:')
                    else:
                        df5["Z"] = df5["B"] - df5["A"]

            elif  value1 == 'A' or value1 == 'A/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:            
                    df5 = pd.concat([dfX1, dfY1, dfA1], axis=1)
                    df5.columns = ['X', 'Y', 'A']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5["Z"] = df5["A"]
    
            elif  value1 == 'B' or value1 == 'B/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                else:            
                    df5 = pd.concat([dfX2, dfY2, dfB1], axis=1)
                    df5.columns = ['X', 'Y', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['B'] = df5['B'].apply(float)
                    df5["Z"] = df5["B"]
            else:
                pass

# save dataset
        elif (index_da == 0):            
            index = index_a + index_b        
            self.canvas_graph.delete('all')       
            id = self.entry_id.get()
            ID = id
            unit = self.entry_unit.get()

            dfX = df2.iloc[0:row_end, 0]
            dfY = df2.iloc[0:row_end, 1]

            value1 = self.variable_cal.get()
            if value1 == 'A-B' or value1 == '(A-B)/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select films for "A" and "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:
                    df5 = pd.concat([dfX, dfY, dfA_cb34, dfB_cb34], axis=1)
                    df5.columns = ['X', 'Y', 'A', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5['B'] = df5['B'].apply(float)
                    
                    if df5['A'].equals(df5['B']):
                        win32api.MessageBox(0, 'You have selected the same film for "A" and "B"', 'Error:')
                    else:
                        df5["Z"] = df5["A"] - df5["B"]
            
            elif value1 == 'B-A' or value1 == '(B-A)/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select films for "A" and "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:            
                    df5 = pd.concat([dfX, dfY, dfA_cb34, dfB_cb34], axis=1)
                    df5.columns = ['X', 'Y', 'A', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5['B'] = df5['B'].apply(float)

                    if df5['A'].equals(df5['B']):
                        win32api.MessageBox(0, 'You have selected the same film for "A" and "B"', 'Error:')
                    else:
                        df5["Z"] = df5["B"] - df5["A"]

            elif  value1 == 'A' or value1 == 'A/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:            
                    df5 = pd.concat([dfX, dfY, dfA_cb34], axis=1)
                    df5.columns = ['X', 'Y', 'A']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5["Z"] = df5["A"]
    
            elif  value1 == 'B' or value1 == 'B/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                else:            
                    df5 = pd.concat([dfX, dfY, dfB_cb34], axis=1)
                    df5.columns = ['X', 'Y', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['B'] = df5['B'].apply(float)
                    df5["Z"] = df5["B"]

# save atlas map
        elif (index_da == 2):
            index = index_a + index_b        
            self.canvas_graph.delete('all')       
            id = self.entry_id.get()
            ID = id
            unit = self.entry_unit.get()
            dfX = df2.iloc[0:row_end, 0]
            dfY = df2.iloc[0:row_end, 1]
            
            value1 = self.variable_cal.get()
            if value1 == 'A-B' or value1 == '(A-B)/t' or value1 == 'B-A' or value1 == '(B-A)/t':
                win32api.MessageBox(0, 'You only have one film. Select "A" or "B"', 'Error:')

            elif  value1 == 'A' or value1 == 'A/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:            
                    df5 = pd.concat([dfX, dfY, dfA_cb56], axis=1)
                    df5.columns = ['X', 'Y', 'A']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5["Z"] = df5["A"]

            elif  value1 == 'B' or value1 == 'B/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                else:            
                    df5 = pd.concat([dfX, dfY, dfB_cb56], axis=1)
                    df5.columns = ['X', 'Y', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['B'] = df5['B'].apply(float)
                    df5["Z"] = df5["B"]
            else:
                pass

# save resmap
        else:
            index = index_a + index_b        
            self.canvas_graph.delete('all')       
            id = self.entry_id.get()
            ID = id
            unit = self.entry_unit.get()
            dfX = df2.iloc[0:row_end, 0]
            dfY = df2.iloc[0:row_end, 1]
            
            value1 = self.variable_cal.get()
            if value1 == 'A-B' or value1 == '(A-B)/t' or value1 == 'B-A' or value1 == '(B-A)/t':
                win32api.MessageBox(0, 'You only have one film. Select "A" or "B"', 'Error:')

            elif  value1 == 'A' or value1 == 'A/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                elif (index == 2):
                    win32api.MessageBox(0, 'You need to select film for "A"', 'Error:')
                else:            
                    df5 = pd.concat([dfX, dfY, dfA_cb78], axis=1)
                    df5.columns = ['X', 'Y', 'A']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['A'] = df5['A'].apply(float)
                    df5["Z"] = df5["A"]

            elif  value1 == 'B' or value1 == 'B/t':
                if (index == 0):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                elif (index == 1):
                    win32api.MessageBox(0, 'You need to select film for "B"', 'Error:')
                else:            
                    df5 = pd.concat([dfX, dfY, dfB_cb78], axis=1)
                    df5.columns = ['X', 'Y', 'B']
                    df5['X'] = df5['X'].apply(float)
                    df5['Y'] = df5['Y'].apply(float)
                    df5['B'] = df5['B'].apply(float)
                    df5["Z"] = df5["B"]
            else:
                pass

# continue save
        z = df5["Z"]
        N_init = df5.shape[0]
        data = [df5["X"], df5["Y"], z]
        headers = ["A", "B", "Z"]
        orignal_df = pd.concat(data, axis=1, keys=headers)

        value = self.radio_sigma.get()
        if (value == 1):
            sigma_var = float(self.entry_sigma.get())
            z_scores = np.abs(stats.zscore(orignal_df))
            df = orignal_df[(z_scores < sigma_var).all(axis=1)]
            N1 = int(orignal_df.shape[0] - df.shape[0])
            N = N_init - N1
            ul_pre = max(df["Z"]) - np.mean(df["Z"])
            ll_pre = np.mean(df["Z"]) -min(df["Z"])
            stdev = np.std(z)
            if ul_pre >  ll_pre:
                sigma = round(ul_pre/stdev, 2)
            else:   
                sigma = round(ll_pre/stdev, 2)
        else:
            N1 = int(self.entry_outlier.get())
            N = N_init - N1
            orignal_df["dist"] = np.abs(orignal_df["Z"]  - np.mean(orignal_df["Z"]))
            sort_df = orignal_df.sort_values(by="dist", ascending=False)
            df = sort_df.iloc[N1:]
            stdev = np.std(z)
            top = max(df["dist"])
            sigma = round(top/stdev, 2)

# create folder and excel
        import xlsxwriter
        from datetime import datetime
        import time
        
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %H-%M-%S")      
        check_dir = os.path.isdir('./Saved Graphs')
        if (check_dir == FALSE):
            os.makedirs('Saved Graphs')
        else:
            pass
        save_path = './Saved Graphs'
        file_name = id + "  " + dt_string + ".xlsx"
        complete_name = os. path. join(save_path, file_name)
        workbook = xlsxwriter.Workbook(complete_name)
        worksheet = workbook.add_worksheet()
        worksheet.set_column('A:A', 12)
        worksheet.set_column('B:B', 12)
        worksheet.set_column('C:C', 12)
        worksheet.set_column('D:D', 12)
        worksheet.set_column('E:E', 12)        

# 2d contour
        global Rate
        global limits_t
        global limits_b

        fig1 = plt.figure(figsize=(2.7, 2.5))
        spec = gridspec.GridSpec(ncols=1, nrows=2, hspace=0.1, height_ratios=[1, 5])

        ax0 = fig1.add_subplot(spec[0, 0])
        ax0.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False) 
        ax0.tick_params(axis='y', which='both', right=False, left=False, labelleft=False) 
        for pos in ['right', 'top', 'bottom', 'left']: plt.gca().spines[pos].set_visible(False)
        
        ax1 = fig1.add_subplot(spec[1, 0])
        ax1.set_aspect('equal', adjustable='box')        
        ax1.set_xlabel('X (mm)')
        ax1.set_ylabel('Y (mm)')    
        ax1.tick_params(axis='both', length=3)
        
        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x2d = df["A"]
            y2d = df["B"]
        else:
            x_original = df["A"]
            y_original = df["B"]        
            x2d = math.cos(radian_theta)*x_original + math.sin(radian_theta)*y_original
            y2d = - math.sin(radian_theta)*x_original + math.cos(radian_theta)*y_original
            
        z2d = df["Z"]
        
        x_grid = np.linspace(np.min(x2d), np.max(x2d), 200)
        y_grid = np.linspace(np.min(y2d), np.max(y2d), 200)
        X2d, Y2d = np.meshgrid(x_grid, y_grid)
        Z2d = griddata((x2d, y2d), z2d, (X2d, Y2d), method="cubic")
        
        Ave = round(np.mean(z2d), 1)
        Std = round(np.std(z2d), 1)
        Std_percent = round(100*Std/abs(Ave), 2)
        Max = round(max(z2d), 1)
        Min = round(min(z2d), 1)
        Nonu = round(0.5*100*(Max-Min)/abs(Ave), 2)
        Range = round(max(z2d) - min(z2d), 1)
        
        value1 = self.variable_cal.get()
        if value1 == '(A-B)/t' or value1 == '(B-A)/t' or value1 == 'A/t' or value1 == 'B/t':
            time_pre = self.entry_run_time.get()
            global time
            time = float(time_pre)
            Rate_pre = abs(Ave)*(60/time)
            Rate = round(Rate_pre, 1)

            ax0.text(0.67, 1.16, 'Rate (' + unit + '/min) =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
            ax0.text(0.94, 1.16, Rate, fontsize=5.7, horizontalalignment='left', verticalalignment='center',
                         transform=ax0.transAxes)
        else:
            pass

        worksheet.write('A1', 'Time')
        worksheet.write('A2', 'ID')
        worksheet.write('A5', 'Mean (' + unit + ')')
        worksheet.write('A6', 'NonU (%)')
        worksheet.write('A7', 'StdD (' + unit + ')')
        worksheet.write('A8', 'StdD / Mean (%)')
        worksheet.write('A9', 'Max (' + unit + ')')
        worksheet.write('A10', 'Min (' + unit + ')')
        worksheet.write('A11', 'Range (' + unit + ')')
        worksheet.write('A12', 'Total Points')
        
        worksheet.write('B1', dt_string)
        worksheet.write('B2', id)
        worksheet.write('B5', abs(Ave))
        worksheet.write('B6', Nonu)
        worksheet.write('B7', Std)
        worksheet.write('B8', Std_percent)
        worksheet.write('B9', Max)
        worksheet.write('B10', Min)
        worksheet.write('B11', Range)
        worksheet.write('B12', N)

        worksheet.write('A14', 'X')
        worksheet.write('B14', 'Y')
        worksheet.write('C14', 'A')
        worksheet.write('D14', 'B')
        worksheet.write('E14', 'Z ')
        worksheet.write_column(14, 0, x2d)
        worksheet.write_column(14, 1, y2d)
        worksheet.write_column(14, 4, z)

        if (index_da == 2):
            df_pre1 = df_pre.replace(np.nan, '')        
            worksheet.write_column(0, 26, df_pre1['a1'])
            worksheet.write_column(0, 27, df_pre1['a2'])
            worksheet.write_column(0, 28, df_pre1['a3'])
            worksheet.write_column(0, 29, df_pre1['a4'])
            worksheet.write_column(0, 30, df_pre1['a5'])
            worksheet.write_column(0, 31, df_pre1['a6'])
        elif (index_da == 3):
            df_pre1 = df_pre.replace(np.nan, '')        
            worksheet.write_column(0, 26, df_pre1.iloc[:, 0])
            worksheet.write_column(0, 27, df_pre1.iloc[:, 1])
            worksheet.write_column(0, 28, df_pre1.iloc[:, 2])
            worksheet.write_column(0, 29, df_pre1.iloc[:, 3])
            worksheet.write_column(0, 30, df_pre1.iloc[:, 4])
            worksheet.write_column(0, 31, df_pre1.iloc[:, 5])
            worksheet.write_column(0, 32, df_pre1.iloc[:, 6])
            worksheet.write_column(0, 33, df_pre1.iloc[:, 7])
            worksheet.write_column(0, 34, df_pre1.iloc[:, 8])
            worksheet.write_column(0, 35, df_pre1.iloc[:, 9])
            worksheet.write_column(0, 36, df_pre1.iloc[:, 10])
            worksheet.write_column(0, 37, df_pre1.iloc[:, 11])
            worksheet.write_column(0, 38, df_pre1.iloc[:, 12])
            worksheet.write_column(0, 39, df_pre1.iloc[:, 13])
            worksheet.write_column(0, 40, df_pre1.iloc[:, 14])
        else:
            pass

        if value1 == 'A' or value1 == 'A/t':
            worksheet.write_column(14, 2, df5['A'])
        elif value1 == 'B' or value1 == 'B/t':
            worksheet.write_column(14, 3, df5['B'])
        else:
            worksheet.write_column(14, 2, df5['A'])
            worksheet.write_column(14, 3, df5['B'])

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_bt_pre = self.limits_entry.get()
            limits_b = float(limits_bt_pre)
            limits_t_pre = self.limits_up.get()
            limits_t = float(limits_t_pre)
            if (limits_t <= limits_b):
                win32api.MessageBox(0, 'Lower limit should be lower than upper limit', 'Error:')
            else:
                pass
            num = int(self.entry_contour.get())
            if (num >150):
                num = 150
                levels = np.linspace(limits_b, limits_t, num)
            else:
                levels = np.linspace(limits_b, limits_t, num)
            cp = plt.contourf(X2d, Y2d, Z2d, levels = levels, cmap=plt.cm.turbo, alpha=0.95)
            cbar1 = fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
            cbar1.ax.locator_params(nbins=6)
        else:
            contour = int(self.entry_contour.get()) 
            if (contour >150):
                contour = 150
            else:
                pass
            cp = plt.contourf(X2d, Y2d, Z2d, contour, cmap=plt.cm.turbo, alpha=0.95)
            cbar1 = fig1.colorbar(cp, ax=ax1, shrink=0.5, orientation="vertical")
            cbar1.ax.locator_params(nbins=6)

        var3 = self.variable3.get()
        if (var3 == "Value"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                if value1 == '(A-B)/t' or value1 == '(B-A)/t' or value1 == 'A/t' or value1 == 'B/t':
                    label = "{:.0f}".format(z2d*(60/time))
                else:
                    label = "{:.0f}".format(z2d)
                plt.annotate(label, (x2d, y2d), textcoords="offset points", xytext=(0, -3), ha='center', 
                            fontsize=4.3, alpha=0.85)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        elif (var3 == "Dot"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.3)
        elif (var3 == "None"):
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                ax1.scatter(x2d, y2d, marker='o', s=1, color='k', alpha=0.0001)
        else:
            for x2d, y2d, z2d in zip(x2d, y2d, z2d):
                label = "{:.0f}".format(z2d)
                if (z2d > Ave):
                    ax1.scatter(x2d, y2d, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                else:
                    ax1.scatter(x2d, y2d, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)
                    
        ax0.text(-0.13, 1.55, 'ID:  ' + id, fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)

        ax0.text(-0.13, 1.16, 'Mean (' + unit + ') =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.27, 1.16, 'NonU (%) =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.09, 1.16, Ave, fontsize=5.7, horizontalalignment='left', verticalalignment='center',
                         transform=ax0.transAxes)
        ax0.text(0.49, 1.16, Nonu, fontsize=5.7, horizontalalignment='left', verticalalignment='center',
                         transform=ax0.transAxes)

        ax0.text(-0.13, 0.82, 'Max (' + unit + ') =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.27, 0.82, 'Min (' + unit + ') =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.67, 0.82, 'Range (' + unit + ') =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.07, 0.82, Max, fontsize=5.7, horizontalalignment='left', verticalalignment='center',
                         transform=ax0.transAxes)
        ax0.text(0.45, 0.82, Min, fontsize=5.7, horizontalalignment='left', verticalalignment='center',
                         transform=ax0.transAxes)
        ax0.text(0.9, 0.82, Range, fontsize=5.7, horizontalalignment='left', verticalalignment='center',
                        transform=ax0.transAxes)
        
        ax0.text(-0.13, 0.48, 'StdD (' + unit + ') =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.27, 0.48, 'StdD/Mean (%) =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.08, 0.48, Std, fontsize=5.7, horizontalalignment='left', verticalalignment='center', 
                         transform=ax0.transAxes)
        ax0.text(0.58, 0.48, Std_percent, fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)

        ax0.text(-0.13, 0.14, 'Total points =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.27, 0.14, 'Points removed =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.67, 0.14, 'Limits (\u00B1\u03C3) =', fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.12, 0.14, N, fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.58, 0.14, N1, fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)
        ax0.text(0.9, 0.14, sigma, fontsize=5.7, horizontalalignment='left',
                         verticalalignment='center', transform=ax0.transAxes)

        plot_id1 = "contour"
        plt.savefig(plot_id1)

        plot_id11 = "contour1"
        plt.savefig(plot_id11)        
        
        img1 = Image.open("contour.png")
        plot_id1_resized = img1.resize((360, 332))
        plot_id1_resized.save("contour_resized.png")     
        img11 = ImageTk.PhotoImage(Image.open("contour_resized.png"))
        self.canvas_graph.create_image(30, 10, anchor="nw", image=img11)
        worksheet.insert_image('G2', 'contour.png', {'x_scale': 1, 'y_scale': 1.05})
        os.remove("contour_resized.png")
        plt.clf()
        plt.close(fig1)

# copy clipboard
        def send_to_clipboard1(clip_type, data):
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(clip_type, data)
            win32clipboard.CloseClipboard()

        image = Image.open("contour1.png")
        output = BytesIO()
        image.convert("RGB").save(output, "BMP")
        data = output.getvalue()[14:]
        output.close()
        send_to_clipboard1(win32clipboard.CF_DIB, data)
        os.remove("contour1.png")
            
# cross_section 1
        fig_crs1 = plt.figure(figsize=(2.3, 1.8))     
        spec_crs1 = gridspec.GridSpec(ncols=1, nrows=1)
        ax_crs1 = fig_crs1.add_subplot(spec_crs1[0, 0])
        ax_crs1.set_xlabel('Cross section (mm)')
        ax_crs1.set_ylabel('Z (' + unit + ')')
        ax_crs1.set(title=id)
        ax_crs1.set_title('Cross-sectional profile, X-X and Y-Y cuts', fontsize=6)

        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x_crs1 = df["A"]
            y_crs1 = df["B"]
        else:
            x_oric1 = df["A"]
            y_oric1 = df["B"]
            x_crs1 = math.cos(radian_theta)*x_oric1+ math.sin(radian_theta)*y_oric1
            y_crs1 = - math.sin(radian_theta)*x_oric1 + math.cos(radian_theta)*y_oric1
        z_crs1 = df["Z"]

        n_crs1 = 80
        x_grid_crs1 = np.linspace(np.min(x_crs1), np.max(x_crs1), n_crs1)
        y_grid_crs1 = np.linspace(np.min(y_crs1), np.max(y_crs1), n_crs1)
        X_crs1, Y_crs1 = np.meshgrid(x_grid_crs1, y_grid_crs1)
        
        Y_crs1 = 0
        Z_crs1 = griddata((x_crs1, y_crs1), z_crs1, (X_crs1, Y_crs1), method="cubic")

        degree_crs1_90 = 90
        radian_crs1_90 = math.radians(degree_crs1_90)     
        x_crs1_90 = math.cos(radian_crs1_90)*x_crs1 + math.sin(radian_crs1_90)*y_crs1
        y_crs1_90 = - math.sin(radian_crs1_90)*x_crs1 + math.cos(radian_crs1_90)*y_crs1

        Y_crs1 = 0
        Z_crs1_90 = griddata((x_crs1_90, y_crs1_90), z_crs1, (X_crs1, Y_crs1), method="cubic")

        ave1 = np.mean(z_crs1)
        count_row = n_crs1
        Std = np.std(z_crs1)
        sigma_var = float(self.entry_sigma.get())

        Z_crs1_1 = [ave1 - 3*Std]*count_row 
        Z_crs1_2 = [ave1]*count_row 
        Z_crs1_3 = [ave1 + 3*Std]*count_row

        ax_crs1.plot(X_crs1, Z_crs1_1, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs1.plot(X_crs1, Z_crs1_2, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs1.plot(X_crs1, Z_crs1_3, '.', color='grey', markersize=0.015, alpha=0.165)

        ax_crs1.plot(X_crs1, Z_crs1, 'o', color='red', markersize=0.01)
        ax_crs1.plot(X_crs1, Z_crs1_90, 'o', color='blue', markersize=0.01)
        
        ax_crs1.legend(("Red: X-X", "Blue: Y-Y"), frameon=False, loc='upper right')
        top_x_crs1 = max(x_crs1)
        btn_x_crs1 = min(x_crs1)
        ax_crs1.set_xlim(btn_x_crs1, top_x_crs1)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_y_pre = self.limits_entry.get()
            btn_y_crs1 = float(limits_y_pre)
            limits_t_pre = self.limits_up.get()
            top_y_crs1 = float(limits_t_pre)
            ax_crs1.set_ylim(btn_y_crs1, top_y_crs1)            
        else:
            top_y_crs1 = Ave + Range*1.8
            btn_y_crs1 = Ave - Range*1.8   
            ax_crs1.set_ylim(btn_y_crs1, top_y_crs1)

        fig_crs1.tight_layout(pad=1)    
        plot_id_crs1 ="cross_section1"
        plt.savefig(plot_id_crs1, bbox_inches='tight')
        img_crs1 = Image.open("cross_section1.png")
        plot_crs1_resized = img_crs1.resize((347, 247))
        plot_crs1_resized.save("cross_resized1.png")     
        img_crs1 = ImageTk.PhotoImage(Image.open("cross_resized1.png"))
        self.canvas_graph.create_image(15, 440, anchor="nw", image=img_crs1)
        worksheet.insert_image('G17', 'cross_section1.png', {'x_scale': 1, 'y_scale': 1})
        os.remove("cross_resized1.png")
        plt.clf()
        plt.close(fig_crs1)

# cross_section 2
        fig_crs2 = plt.figure(figsize=(2.3, 1.8))     
        spec_crs2 = gridspec.GridSpec(ncols=1, nrows=1)
        ax_crs2 = fig_crs2.add_subplot(spec_crs2[0, 0])
        ax_crs2.set_xlabel('Cross section (mm)')
        ax_crs2.set_ylabel('Z (' + unit + ')')
        ax_crs2.set(title=id)
        ax_crs2.set_title('Cross-sectional profile, 3 X-X cuts', fontsize=6)

        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x_crs2 = df["A"]
            y_crs2 = df["B"]
        else:
            x_oric2 = df["A"]
            y_oric2 = df["B"]
            x_crs2 = math.cos(radian_theta)*x_oric2+ math.sin(radian_theta)*y_oric2
            y_crs2 = - math.sin(radian_theta)*x_oric2 + math.cos(radian_theta)*y_oric2
        z_crs2 = df["Z"]

        x_grid_crs2 = np.linspace(np.min(x_crs2), np.max(x_crs2), 80)
        y_grid_crs2 = np.linspace(np.min(y_crs2), np.max(y_crs2), 80)
        X_crs2, Y_crs2 = np.meshgrid(x_grid_crs2, y_grid_crs2)

        Y_max = max(y_crs2)
        Y_min = min(y_crs2)
        ran_crs2 = (max(y_crs2) - min(y_crs2))/4
        Y_cros1 = - ran_crs2
        Y_cros3 = 0
        Y_cros5 = ran_crs2

        Z_crs2_1 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros1), method="cubic")
        Z_crs2_3 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros3), method="cubic")
        Z_crs2_5 = griddata((x_crs2, y_crs2), z_crs2, (X_crs2, Y_cros5), method="cubic")

        ax_crs2.plot(X_crs2, Z_crs1_1, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs2.plot(X_crs2, Z_crs1_2, '.', color='grey', markersize=0.015, alpha=0.165)
        ax_crs2.plot(X_crs2, Z_crs1_3, '.', color='grey', markersize=0.015, alpha=0.165)
        
        ax_crs2.plot(X_crs2, Z_crs2_5, 'o', color='orange', markersize=0.01)      
        ax_crs2.plot(X_crs2, Z_crs2_3, 'o', color='red', markersize=0.01)
        ax_crs2.plot(X_crs2, Z_crs2_1, 'o', color='green', markersize=0.01)
           
        ax_crs2.legend(("Green: 1st X-X", "Red: 2nd X-X", "Yellow: 3rd X-X"), frameon=False,
                            loc='upper right')
        top_x_crs2 = max(x_crs2)
        btn_x_crs2 = min(x_crs2)
        ax_crs2.set_xlim(btn_x_crs2, top_x_crs2)

        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_y_pre = self.limits_entry.get()
            btn_y_crs2 = float(limits_y_pre)
            limits_t_pre = self.limits_up.get()
            top_y_crs2 = float(limits_t_pre)
            ax_crs2.set_ylim(btn_y_crs2, top_y_crs2)            
        else:
            top_y_crs2 = Ave + Range*1.8
            btn_y_crs2 = Ave - Range*1.8   
            ax_crs2.set_ylim(btn_y_crs2, top_y_crs2)

        fig_crs2.tight_layout(pad=1)    
        plot_id_crs2 ="cross_section2"
        plt.savefig(plot_id_crs2, bbox_inches='tight')
        img_crs2 = Image.open("cross_section2.png")
        plot_crs2_resized = img_crs2.resize((347, 247))
        plot_crs2_resized.save("cross_resized2.png")     
        img_crs2 = ImageTk.PhotoImage(Image.open("cross_resized2.png"))
        self.canvas_graph.create_image(400, 440, anchor="nw", image=img_crs2)
        worksheet.insert_image('L17', 'cross_section2.png', {'x_scale': 1, 'y_scale': 1})
        os.remove("cross_resized2.png")
        plt.clf()
        plt.close(fig_crs2)

# 3d_contour
        fig3d = plt.figure(figsize=(2.5, 2.5))
        ax3d = plt.axes(projection="3d")
        ax3d.view_init(30, 240)
        degree_theta = float(self.rotation_entry.get())
        radian_theta = math.radians(degree_theta)
        if (degree_theta == 0):
            x3d = df["A"]
            y3d = df["B"]
        else:
            degree_theta = float(self.rotation_entry.get())
            radian_theta = math.radians(degree_theta)
            x_orig = df["A"]
            y_orig = df["B"]
            x3d = math.cos(radian_theta)*x_orig + math.sin(radian_theta)*y_orig
            y3d = - math.sin(radian_theta)*x_orig + math.cos(radian_theta)*y_orig
            
        z3d = df["Z"] 
        x3d_grid = np.linspace(np.min(x3d), np.max(x3d), 120)
        y3d_grid = np.linspace(np.min(y3d), np.max(y3d), 120)
        X3d, Y3d = np.meshgrid(x3d_grid, y3d_grid)
        Z3d = griddata((x3d, y3d), z3d, (X3d, Y3d), method="cubic")
        ax3d.contour3D(X3d, Y3d, Z3d, 300, cmap='turbo', alpha=1, antialiased=False)
 
        check_box = self.var_limits.get() 
        if (check_box == 1):
            limits_z_pre = self.limits_entry.get()
            btn_limitz3d = float(limits_z_pre)
            limits_t_pre = self.limits_up.get()
            top_limitz3d = float(limits_t_pre)
            ax3d.set_zlim(btn_limitz3d, top_limitz3d)          
        else:
            top_limitz3d = np.mean(z3d) + (max(z3d)-min(z3d))*1.5
            btn_limitz3d = np.mean(z3d) - (max(z3d)-min(z3d))*4.5
            ax3d.set_zlim(btn_limitz3d, top_limitz3d)

        cset = ax3d.contour(X3d, Y3d, Z3d, 10, zdir='z', offset=btn_limitz3d, linewidths=1,
                            cmap=cm.turbo, alpha=0.9)
        ax3d.set_xlabel('X (mm)')
        ax3d.set_ylabel('Y (mm)')
        ax3d.set_zlabel(unit)
        ax3d.set_title(' Original 3d map: \n Mean = %.1f unit, NonU = %.2f pct' %(Ave, Nonu),
                           loc='left', fontsize=6)
        
        fig3d.tight_layout(pad=1)    
        plot_id3d ="3d"
        plt.savefig(plot_id3d, bbox_inches='tight')
        img3d = Image.open("3d.png")
        plot_id3d_resized = img3d.resize((356, 356))
        plot_id3d_resized.save("3d_resized.png")     
        img33d = ImageTk.PhotoImage(Image.open("3d_resized.png"))
        self.canvas_graph.create_image(420, 10, anchor="nw", image=img33d)
        worksheet.insert_image('L1', '3d.png', {'x_scale': 1, 'y_scale': 1})
        os.remove("3d_resized.png")
        plt.clf()
        plt.close(fig3d)

# 3d animated
        from matplotlib.animation import FuncAnimation

        check_box3 = self.var_animated.get() 
        if (check_box3 == 1):
            fig_3d_rot = plt.figure()
            fig_3d_rot.set_size_inches(2, 2)
            ax_3d_rot = plt.axes(projection='3d')
            ax_3d_rot.set_facecolor((1, 1, 1))

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_up.get()
                top_limitz3d_animated = float(limits_z_pre)             
                limits_z_pre2 = self.limits_entry.get()
                btn_limitz3d_animated = float(limits_z_pre2)
                ax_3d_rot.set_zlim(btn_limitz3d, top_limitz3d)          
            else:
                top_limitz3d_animated = np.mean(z3d) + (max(z3d)-min(z3d))*1.85
                btn_limitz3d_animated = np.mean(z3d) - (max(z3d)-min(z3d))*1.85
                ax_3d_rot.set_zlim(btn_limitz3d_animated, top_limitz3d_animated)

            ax_3d_rot.set_xlabel('X', fontsize=4)
            ax_3d_rot.set_ylabel('Y', fontsize=4)
            ax_3d_rot.set_zlabel('Z', fontsize=4)
            ax_3d_rot.set_title(id, fontsize=5)

            ax_3d_rot.plot_surface(X3d, Y3d, Z3d, rstride=6, cstride=6, cmap='viridis', alpha=0.85,
                            edgecolor='none')
            plt.axis('on')

            def update(i, fig_3d_rot, ax_3d_rot):
                ax_3d_rot.view_init(elev=30., azim=i)
                return fig_3d_rot, ax_3d_rot

            anim = FuncAnimation(fig_3d_rot, update, frames=np.arange(0, 360, 1), repeat=True,
                            fargs=(fig_3d_rot, ax_3d_rot), interval=100)
            anim.save('rgb_cube.gif', dpi=15, writer='pillow')
            plt.pause(0.1)
            os.remove('rgb_cube.gif')
            
            #for website
            #anim.save('rgb_cube.gif', dpi=600, writer='pillow')             
            
        else:
            pass
            
# plot all
        check_box2 = self.var_decom.get() 
        if (check_box2 == 1):
            
# scatter1
            fig5 = plt.figure(figsize=(2.3, 1.7))
            spec5 = gridspec.GridSpec(ncols=1, nrows=1)

            df["radius (mm)"] = np.sqrt(df["A"]**2 + df["B"]**2)
            df["theta_pre (degree)"] = np.arctan2(df["B"], df["A"])*180/np.pi
            df["theta (degree)"] = df["theta_pre (degree)"].apply(lambda x: x + 360 if x < 0 else x)

            ax5 = fig5.add_subplot(spec5[0, 0])       
            a1 = 1
            b1 = 0.1
            df["u"] = a1*df["radius (mm)"] + b1*df["theta (degree)"]
            df = df.sort_values(by="u")
            df["Point"] = range(1, 1 + len(df["u"]))
            x1 = df["Point"] 
            y1 = df["Z"]

            ave1 = np.mean(y1)
            count_row = df.shape[0]
            Std = np.std(y1)
            sigma_var = float(self.entry_sigma.get())

            y2 = [ave1 - 3*Std]*count_row 
            y3 = [ave1]*count_row 
            y4 = [ave1 + 3*Std]*count_row

            ax5.plot(x1, y1, 'o', markersize=1.5, alpha=0.7)
            ax5.plot(x1, y2, '-', markersize=0.001, c='grey', alpha=0.15)
            ax5.plot(x1, y3, '-', markersize=0.001, c='grey', alpha=0.15)
            ax5.plot(x1, y4, '-', markersize=0.001, c='grey', alpha=0.15)

            ax5.set_xlabel('Measurement point', fontsize=5)
            ax5.set_ylabel('Z (' + unit + ')', fontsize=5)
            ax5.set(title = 'As radius increasing')

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit1_bt_pre = self.limits_entry.get()
                btn_limit1 = float(limit1_bt_pre)
                limits_t_pre = self.limits_up.get()
                top_limit1 = float(limits_t_pre)
                ax5.set_ylim(btn_limit1, top_limit1)
            else:
                top_limit1 = Ave + Range*1.8
                btn_limit1 = Ave - Range*1.8
                ax5.set_ylim(btn_limit1, top_limit1)

            fig5.tight_layout(pad=1)
            plot_id5 ="scatter1"
            plt.savefig(plot_id5, bbox_inches='tight')
            img5 = Image.open("scatter1.png")
            plot_id5_resized = img5.resize((352, 247))
            plot_id5_resized.save("scatter1_resized.png")      
            img5 = ImageTk.PhotoImage(Image.open("scatter1_resized.png")) 
            self.canvas_graph.create_image(15, 710, anchor="nw", image=img5)
            worksheet.insert_image('G28', 'scatter1.png', {'x_scale': 1, 'y_scale': 1})
            os.remove("scatter1_resized.png")
            plt.clf()
            plt.close(fig5)

# scatter2
            fig6 = plt.figure(figsize=(2.3, 1.7))
            spec6 = gridspec.GridSpec(ncols=1, nrows=1)

            df["radius (mm)"] = np.sqrt(df["A"]**2 + df["B"]**2)
            df["theta_pre (degree)"] = np.arctan2(df["B"], df["A"])*180/np.pi
            df["theta (degree)"] = df["theta_pre (degree)"].apply(lambda x: x + 360 if x < 0 else x)

            ax6 = fig6.add_subplot(spec6[0, 0]) 
            a2 = 0.1
            b2 = 1
            df["v"] = a2*df["radius (mm)"] + b2*df["theta (degree)"]    
            df = df.sort_values(by="v")
            df["Point"] = range(1, 1 + len(df["v"]))
            x2 = df["Point"] 
            y2 = df["Z"] 

            ave1 = np.mean(y1)
            count_row = df.shape[0]
            Std = np.std(y1)
            sigma_var = float(self.entry_sigma.get())

            y2_2 = [ave1 - 3*Std]*count_row 
            y2_3 = [ave1]*count_row 
            y2_4 = [ave1 + 3*Std]*count_row 

            ax6.plot(x2, y2, 'o', markersize=1.5, c='red', alpha=0.7)
            ax6.plot(x2, y2_2, '-', markersize=0.001, c='grey', alpha=0.15)
            ax6.plot(x2, y2_3, '-', markersize=0.001, c='grey', alpha=0.15)
            ax6.plot(x2, y2_4, '-', markersize=0.001, c='grey', alpha=0.15)

            ax6.set_xlabel('Measurement point', fontsize=5)
            ax6.set_ylabel('Z (' + unit + ')', fontsize=5)
            ax6.set(title='As theta increasing')

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit2_bt_pre = self.limits_entry.get()
                btn_limit2 = float(limit2_bt_pre)
                limits_t_pre = self.limits_up.get()
                top_limit2 = float(limits_t_pre)
                ax6.set_ylim(btn_limit2, top_limit2)
            else:
                top_limit2 = Ave + Range*1.8
                btn_limit2 = Ave - Range*1.8
                ax6.set_ylim(btn_limit2, top_limit2)

            fig6.tight_layout(pad=1)
            plot_id6 ="scatter2"
            plt.savefig(plot_id6, bbox_inches='tight')
            img6 = Image.open("scatter2.png")
            plot_id6_resized = img6.resize((352, 247))
            plot_id6_resized.save("scatter2_resized.png")      
            img6 = ImageTk.PhotoImage(Image.open("scatter2_resized.png")) 
            self.canvas_graph.create_image(400, 710, anchor="nw", image=img6)
            worksheet.insert_image('L28', 'scatter2.png', {'x_scale': 1, 'y_scale': 1})
            os.remove("scatter2_resized.png")
            plt.clf()
            plt.close(fig6)

# histogram
            fig10 = plt.figure(figsize=(2.5, 1.9))
            spec10 = gridspec.GridSpec(ncols=1, nrows=1)       
            ax10 = fig10.add_subplot(spec10[0, 0])
            ax10.set_xlabel('Z (' + unit + ')', fontsize=6)
            ax10.set_ylabel('Counts', fontsize=6)
            ax10.set_title('Histogram', fontsize=6.5)
            
            z_hist = df["Z"]
            bins = round(len(z_hist)/5)

            z1 = ave1 - 3*Std
            z2 = ave1
            z3 = ave1 + 3*Std

            ax10.hist(z_hist, bins, color='darkblue', alpha=0.5)
            plt.axvline(z1, color='grey', linestyle=':', linewidth=0.5)
            plt.axvline(z2, color='grey', linestyle=':', linewidth=0.5)
            plt.axvline(z3, color='grey', linestyle=':', linewidth=0.5)

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit_pre = self.limits_entry.get()
                btn_limit = float(limit_pre)
                limits_t_pre = self.limits_up.get()
                top_limit = float(limits_t_pre)
                ax10.set_xlim((btn_limit, top_limit))
            else:
                btn_limit = Ave - Range*1.3
                top_limit = Ave + Range*1.3
                ax10.set_xlim((btn_limit, top_limit))

            fig10.tight_layout(pad=1)
            plot_id10 ="hist"
            plt.savefig(plot_id10, bbox_inches='tight')
            img10 = Image.open("hist.png")
            plot_id10_resized = img10.resize((332, 242))
            plot_id10_resized.save("hist_resized.png")      
            img10 = ImageTk.PhotoImage(Image.open("hist_resized.png")) 
            self.canvas_graph.create_image(37, 980, anchor="nw", image=img10)
            worksheet.insert_image('G39', 'hist.png', {'x_scale': 0.9, 'y_scale': 0.9})
            os.remove("hist_resized.png")
            plt.clf()
            plt.close(fig10)

# cumulative distribution
            fig_cdf = plt.figure(figsize=(2.5, 1.9))
            spec_cdf = gridspec.GridSpec(ncols=1, nrows=1)       
            ax_cdf = fig_cdf.add_subplot(spec_cdf[0, 0])
            ax_cdf.set_xlabel('Z (' + unit + ')', fontsize=6)
            ax_cdf.set_ylabel('Percent', fontsize=6)
            ax_cdf.set_title('Cumulative distribution', fontsize=6.5)

            x_cdf_pre = df["Z"]
            x_cdf = np.sort(x_cdf_pre)
            y_cdf = 100*np.arange(len(x_cdf))/float(len(x_cdf))

            ave = np.mean(x_cdf_pre)
            count_row = df.shape[0]
            Std = np.std(x_cdf_pre)
            sigma_var = float(self.entry_sigma.get())

            x_cdf_2= [ave - 3*Std]*count_row 
            x_cdf_3 = [ave]*count_row 
            x_cdf_4 = [ave + 3*Std]*count_row 

            ax_cdf.plot(x_cdf, y_cdf, 'o', markersize=1.8, color='m', alpha=0.7)
            ax_cdf.plot(x_cdf_2, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)
            ax_cdf.plot(x_cdf_3, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)
            ax_cdf.plot(x_cdf_4, y_cdf, '-', markersize=0.001, c='grey', alpha=0.15)

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limit_pre = self.limits_entry.get()
                btn_limit = float(limit_pre)
                limits_t_pre = self.limits_up.get()
                top_limit = float(limits_t_pre)
                ax_cdf.set_xlim((btn_limit, top_limit))
            else:
                top_cdf = Ave + Range*1.3
                btn_cdf = Ave - Range*1.3
                ax_cdf.set_xlim(btn_cdf, top_cdf)

            fig_cdf.tight_layout(pad=1)
            plot_id_cdf ="cdf"
            plt.savefig(plot_id_cdf, bbox_inches='tight')
            img_cdf = Image.open("cdf.png")
            plot_id_cdf_resized = img_cdf.resize((335, 242))
            plot_id_cdf_resized.save("cdf_resized.png")      
            img_cdf = ImageTk.PhotoImage(Image.open("cdf_resized.png")) 
            self.canvas_graph.create_image(422, 980, anchor="nw", image=img_cdf)
            worksheet.insert_image('L39', 'cdf.png', {'x_scale': 0.9, 'y_scale': 0.9})
            os.remove("cdf_resized.png")
            plt.clf()
            plt.close(fig_cdf)

# decompose_tilt_corrected (3d) #1
            fig_slp1 = plt.figure(figsize=(2.5, 2.5))
            ax_slp1 = plt.axes(projection="3d")
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
                                    
            if (degree_theta == 0):
                df["x_tc"] = df["A"]
                df["y_tc"] = df["B"]
            else:
                x_orir = df["A"]
                y_orir = df["B"]
                df["x_tc"] = math.cos(radian_rotation)*x_orir + math.sin(radian_rotation)*y_orir
                df["y_tc"] = -math.sin(radian_rotation)*x_orir + math.cos(radian_rotation)*y_orir

            x_slp1 = df["x_tc"]
            y_slp1 = df["y_tc"]
            z_slp1 = df["Z"]

            x_tmp = df[['x_tc', 'y_tc']].values.reshape(-1, 2)
            max_rows = max(len(x_slp1), len(y_slp1))
            x_ones = np.ones([max_rows, 1])
            x = np.append(x_ones, x_tmp, axis=1)
            y_tmp = z_slp1.tolist()
            y = np.array(y_tmp)

            def mse(coef, x, y):
                return np.mean((np.dot(x, coef) - y)**2)/2
            
            def gradients(coef, x, y):
                return np.mean(x.transpose()*(np.dot(x, coef) - y), axis=1)
            
            def multilinear_regression(coef, x, y, lr, b1=0.9, b2=0.999, epsilon=1e-8):
                prev_error = 0
                m_coef = np.zeros(coef.shape)
                v_coef = np.zeros(coef.shape)
                moment_m_coef = np.zeros(coef.shape)
                moment_v_coef = np.zeros(coef.shape)
                t = 0
                while True:
                    error = mse(coef, x, y)
                    if abs(error - prev_error) <= epsilon:
                        break
                    prev_error = error
                    grad = gradients(coef, x, y)
                    t += 1
                    m_coef = b1*m_coef + (1-b1)*grad
                    v_coef = b2*v_coef + (1-b2)*grad**2
                    moment_m_coef = m_coef / (1-b1**t)
                    moment_v_coef = v_coef / (1-b2**t)
                    delta = ((lr / moment_v_coef**0.5 + 1e-8)*(b1*moment_m_coef + (1-b1)*grad/(1-b1**t)))
                    coef = np.subtract(coef, delta)
                return coef
 
            coef = np.array([0, 0, 0])
            c = multilinear_regression(coef, x, y, 1e-1)
        
            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                btn_limitz_slp1 = float(limits_z_pre)
                limits_t_pre = self.limits_up.get()
                top_limitz_slp1 = float(limits_t_pre)
            else:
                ax_slp1.set_zlim(btn_limitz3d, top_limitz3d)
                btn_limitz_slp1 = np.mean(z_slp1) - (max(z_slp1)-min(z_slp1))*4.5
                top_limitz_slp1 = np.mean(z_slp1) + (max(z_slp1)-min(z_slp1))*1.5

            x_slp1_grid = np.linspace(np.min(x_slp1), np.max(x_slp1), 120)
            y_slp1_grid = np.linspace(np.min(y_slp1), np.max(y_slp1), 120)
            X_slp1, Y_slp1 = np.meshgrid(x_slp1_grid, y_slp1_grid)
            Z_slp1 = griddata((x_slp1, y_slp1), z_slp1, (X_slp1, Y_slp1), method="cubic")
            zz_slp1 = c[1]*x_slp1 + c[2]*y_slp1 + c[0]
            ZZ_slp1 = griddata((x_slp1, y_slp1), zz_slp1, (X_slp1, Y_slp1), method="linear")
            
            ax_slp1.contour3D(X_slp1, Y_slp1, Z_slp1 - ZZ_slp1 + c[0], 300, cmap='turbo', alpha=1,
                            antialiased=False)
            
            cset = ax_slp1.contour(X_slp1, Y_slp1, Z_slp1 - ZZ_slp1 + c[0], 7, zdir='z',
                            offset=btn_limitz_slp1, linewidths=1, cmap=cm.turbo, alpha=0.9)

            z_after1 = z_slp1 - zz_slp1 + c[0]
            ave_slp1 = round(np.mean(z_after1), 1)
            max_slp1 = round(max(z_after1))
            min_slp1 = round(min(z_after1))
            nonu_slp1 = round(0.5*100*(max_slp1 - min_slp1)/abs(ave_slp1), 2)

            ax_slp1.set_title(' After tilting is corrected: \n Mean = %.1f unit, NonU = %.2f pct'
                            %(ave_slp1, nonu_slp1), loc='left', fontsize=5.5)
            ax_slp1.set_zlim(btn_limitz_slp1, top_limitz_slp1)
            ax_slp1.view_init(30, 240)
            ax_slp1.set_xlabel('X (mm)')
            ax_slp1.set_ylabel('Y (mm)')
            ax_slp1.set_zlabel(unit)

            plot_slp1 ="slope1"
            plt.savefig(plot_slp1, bbox_inches='tight')
            img_slp1 = Image.open("slope1.png")
            plot_slp1_resized = img_slp1.resize((356, 366))
            plot_slp1_resized.save("slope1_resized.png")     
            img_slp1 = ImageTk.PhotoImage(Image.open("slope1_resized.png"))
            self.canvas_graph.create_image(807, 10, anchor="nw", image=img_slp1)
            worksheet.insert_image('Q1', 'slope1.png', {'x_scale': 1.1, 'y_scale': 1.1})
            os.remove("slope1_resized.png")
            plt.clf()
            plt.close(fig_slp1)

# decompose_tilt_corrected (2d) #1
            fig_slp2d = plt.figure(figsize=(2.0, 2.0))
            spec_slp2d = gridspec.GridSpec(ncols=1, nrows=1)
            ax_slp2d = fig_slp2d.add_subplot(spec_slp2d[0, 0])

            X_slp2d = X_slp1
            Y_slp2d = Y_slp1
            Z_slp2d = Z_slp1
            ZZ_slp2d = ZZ_slp1
            
            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_bt_pre = self.limits_entry.get()
                limits_b = float(limits_bt_pre)
                limits_t_pre = self.limits_up.get()
                limits_t = float(limits_t_pre)
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_slp2d, Y_slp2d, Z_slp2d - ZZ_slp2d + c[0], levels = levels,
                            cmap=plt.cm.turbo, alpha=0.95)
            else:
                limits_b = Min*0.97
                limits_t = Max*1.03
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_slp2d, Y_slp2d, Z_slp2d - ZZ_slp2d + c[0], levels = levels,
                            cmap=plt.cm.turbo, alpha=0.95)

            ax_slp2d.set_aspect('equal', adjustable='box')
            ax_slp2d.set_title(' \n After tilting is corrected: \n Mean = %.1f unit, NonU = %.2f pct'
                            %(ave_slp1, nonu_slp1), loc='left', fontsize=5.5)

            x_slp2d = x_slp1 
            y_slp2d = y_slp1
            z6 = z_after1
            
            var3 = self.variable3.get()
            if (var3 == "Value"):
                for x_slp2d, y_slp2d, z6 in zip(x_slp2d, y_slp2d, z6):
                    if value1 == '(A-B)/t' or value1 == '(B-A)/t' or value1 == 'A/t' or value1 == 'B/t':
                        label = "{:.0f}".format(z6*(60/time))
                    else:
                        label = "{:.0f}".format(z6)
                    plt.annotate(label, (x_slp2d, y_slp2d), textcoords="offset points", xytext=(0, -3), ha='center',
                                 fontsize=4.3, alpha=0.85)
                    ax_slp2d.scatter(x_slp2d, y_slp2d, marker='o', s=1, color='k', alpha=0.0001)
            elif (var3 == "Dot"):
                for x_slp2d, y_slp2d, z6 in zip(x_slp2d, y_slp2d, z6):
                    ax_slp2d.scatter(x_slp2d, y_slp2d, marker='o', s=1, color='k', alpha=0.3)
            elif (var3 == "Blank"):
                for x_slp2d, y_slp2d, z6 in zip(x_slp2d, y_slp2d, z6):
                    ax_slp2d.scatter(x_slp2d, y_slp2d, marker='o', s=1, color='k', alpha=0.0001)
            else:
                for x_slp2d, y_slp2d, z6 in zip(x_slp2d, y_slp2d, z6):
                    label = "{:.0f}".format(z6)
                    if (z6 > Ave):
                        ax_slp2d.scatter(x_slp2d, y_slp2d, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                    else:
                        ax_slp2d.scatter(x_slp2d, y_slp2d, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)
            
            ax_slp2d.set_xlabel('X (mm)')
            ax_slp2d.set_ylabel('Y (mm)')
            ax_slp2d.tick_params(axis='both', length=3)
            fig_slp2d.tight_layout(pad=0.1)
            plot_slp2d = "contour_slp2d"
            plt.savefig(plot_slp2d, bbox_inches='tight')        
            img2d = Image.open("contour_slp2d.png")
            plot_slp2d_resized = img2d.resize((288, 312))
            plot_slp2d_resized.save("contour2d_resized.png")     
            img2d = ImageTk.PhotoImage(Image.open("contour2d_resized.png"))
            self.canvas_graph.create_image(857, 780, anchor="nw", image=img2d)
            worksheet.insert_image('Q31', 'contour_slp2d.png', {'x_scale': 1.08, 'y_scale': 1.12})
            os.remove("contour2d_resized.png")
            plt.clf()
            plt.close(fig_slp2d)
        
# decompose_slope (3d) #3
            fig_slp2 = plt.figure(figsize=(2.5, 2.5))
            ax_slp2 = plt.axes(projection="3d")
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
            
            if (degree_theta == 0):
                x_slp2 = df["A"]
                y_slp2 = df["B"]
            else:
                x_orir = df["A"]
                y_orir = df["B"]
                x_slp2 = math.cos(radian_rotation)*x_orir + math.sin(radian_rotation)*y_orir
                y_slp2 = -math.sin(radian_rotation)*x_orir + math.cos(radian_rotation)*y_orir

            z_slp2 = df["Z"]
            zz_slp2 = c[1]*x_slp2 + c[2]*y_slp2 + c[0]

            actual = y_tmp
            predict = zz_slp2
            correlation_matrix = np.corrcoef(actual, predict)
            corr = correlation_matrix[0, 1]
            R_sq1 = corr**2
            
            v1 = (c[1], c[2], -10**(-7))
            v2 = (0, 0, 10**(-7))
            def unit_vector(vector):
                return vector / np.linalg.norm(vector)
            def angle_between(v1, v2):
                v1_u = unit_vector(v1)
                v2_u = unit_vector(v2)
                return np.arccos(np.clip(np.dot(v1_u, v2_u), -1.0, 1.0))
            angle_deg = np.round(180 - math.degrees(angle_between(v1, v2)), decimals=9, out=None)
            angle_deg2 = (np.round(90 - angle_deg, decimals=9, out=None))*10**6

            u1 = (1, 0)
            u2 = (c[1], c[2])
            def unit_vector(vector):
                return vector / np.linalg.norm(vector)
            def angle_between1(u1, u2):
                u11 = unit_vector(u1)
                u22 = unit_vector(u2)
                return np.arccos(np.clip(np.dot(u11, u22), -1.0, 1.0))
            non_acute = np.round(180 - math.degrees(angle_between1(u1, u2)), decimals=1, out=None)

            x_slp2_grid = np.linspace(np.min(x_slp2), np.max(x_slp2), 120)
            y_slp2_grid = np.linspace(np.min(y_slp2), np.max(y_slp2), 120)
            X_slp2, Y_slp2 = np.meshgrid(x_slp2_grid, y_slp2_grid)
            Z_slp2 = griddata((x_slp2, y_slp2), zz_slp2, (X_slp2, Y_slp2), method="linear")
                
            btn_limitz_slp2 = btn_limitz3d
            top_limitz_slp2 = top_limitz3d
            Z_slp3 = Z_slp2 - c[0] + np.mean(z3d)
            
            ax_slp2.contour3D(X_slp2, Y_slp2, Z_slp3, 100, cmap='turbo', alpha=0.55, antialiased=False)
            cset = ax_slp2.contour(X_slp2, Y_slp2, Z_slp2, 7, zdir='z', offset=btn_limitz_slp2, linewidths=1,
                            cmap=cm.turbo, alpha=0.9)

            ax_slp2.set_zlim(btn_limitz_slp2, top_limitz_slp2)            
            ax_slp2.view_init(30, 240)
            ax_slp2.set_xlabel('X (mm)')
            ax_slp2.set_ylabel('Y (mm)')
            ax_slp2.set_zlabel(unit)

            ax_slp2.set_title(
                ' Abstracted tilting: \n z = (%.2f) x + (%.2f) y + (%.2f), R$^2$ = %.2f \n Tilt = %.2f E-6°, Skew = %.1f°'
                % (c[1], c[2], c[0], R_sq1, angle_deg2, non_acute), loc='left', fontsize=5.5)

            fig_slp2.tight_layout(pad=1)
            plot_slp2 ="slope2"
            plt.savefig(plot_slp2, bbox_inches='tight')
            img_slp2 = Image.open("slope2.png")
            plot_slp2_resized = img_slp2.resize((356, 386))
            plot_slp2_resized.save("slope2_resized.png")     
            img_slp2 = ImageTk.PhotoImage(Image.open("slope2_resized.png"))
            self.canvas_graph.create_image(814, 380, anchor="nw", image=img_slp2)
            worksheet.insert_image('Q16', 'slope2.png', {'x_scale': 1.05, 'y_scale': 1.05})
            os.remove("slope2_resized.png")
            plt.clf()
            plt.close(fig_slp2)

# decompose_slope (2d) #3
            fig_skew = plt.figure(figsize=(2.22, 2.22))
            spec_skew = gridspec.GridSpec(ncols=1, nrows=1)
            ax_skew = fig_skew.add_subplot(spec_skew[0, 0])

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_bt_pre = self.limits_entry.get()
                limits_b = float(limits_bt_pre)
                limits_t_pre = self.limits_up.get()
                limits_t = float(limits_t_pre)
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_slp2, Y_slp2, Z_slp2, levels = levels, cmap=plt.cm.turbo, alpha=0.95)
            else:
                limits_b = Min
                limits_t = Max
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_slp2, Y_slp2, Z_slp2, levels = levels, cmap=plt.cm.turbo, alpha=0.95)

            x_skew = x_slp1 
            y_skew = y_slp1
            z_skew = zz_slp2 - c[0]
            df_z = z_skew.to_frame(name="ht")
            range_skew = round((df_z["ht"].max() - df_z["ht"].min()), 1)
            nonu_skew = round(0.5*100*(df_z["ht"].max() - df_z["ht"].min())/abs(Ave), 2)
            
            var3 = self.variable3.get()
            if (var3 == "Value"):
                for x_skew, y_skew, z_skew in zip(x_skew, y_skew, z_skew):
                    if value1 == '(A-B)/t' or value1 == '(B-A)/t' or value1 == 'A/t' or value1 == 'B/t':
                        label = "{:.0f}".format(z_skew*(60/time))
                    else:
                        label = "{:.0f}".format(z_skew)
                    plt.annotate(label, (x_skew, y_skew), textcoords="offset points", xytext=(0, -3), ha='center',
                                 fontsize=4.3, alpha=0.85)
                    ax_skew.scatter(x_skew, y_skew, marker='o', s=1, color='k', alpha=0.0001)
            elif (var3 == "Dot"):
                for x_skew, y_skew, z_skew in zip(x_skew, y_skew, z_skew):
                    ax_skew.scatter(x_skew, y_skew, marker='o', s=1, color='k', alpha=0.3)
            elif (var3 == "Blank"):
                for x_skew, y_skew, z_skew in zip(x_skew, y_skew, z_skew):
                    ax_skew.scatter(x_skew, y_skew, marker='o', s=1, color='k', alpha=0.0001)
            else:
                for x_skew, y_skew, z_skew in zip(x_skew, y_skew, z_skew):
                    label = "{:.0f}".format(z_skew)
                    if (z_skew > 0):
                        ax_skew.scatter(x_skew, y_skew, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                    else:
                        ax_skew.scatter(x_skew, y_skew, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)

            ax_skew.set_aspect('equal', adjustable='box')
            ax_skew.set_title(' Abstracted tilting: \n Tilt = %.2f E-6°, Skew = %.1f° \n Range = %.1f unit, NonU = %.2f pct'
                            %(angle_deg2, non_acute, range_skew, nonu_skew), loc='left', fontsize=6)
            ax_skew.set_xlabel('X (mm)')
            ax_skew.set_ylabel('Y (mm)')
            ax_skew.tick_params(axis='both', length=3)

            fig_skew.tight_layout(pad=0.1)
            plot_skew = "contour_skew"
            plt.savefig(plot_skew, bbox_inches='tight')        
            img_skew = Image.open("contour_skew.png")
            skew_resized = img_skew.resize((275, 297))
            skew_resized.save("skew_resized.png")     
            img_skew = ImageTk.PhotoImage(Image.open("skew_resized.png"))
            self.canvas_graph.create_image(862, 1110, anchor="nw", image=img_skew)
            worksheet.insert_image('Q43', 'contour_skew.png', {'x_scale': 0.95, 'y_scale': 0.97})
            os.remove("skew_resized.png")
            plt.clf()
            plt.close(fig_skew)

# decompose_radius (3d) #4
            from scipy.optimize import curve_fit
            x_rad = df["A"]
            y_rad = df["B"]
            z_rad = df["Z"]
            x = df["radius (mm)"]
            y = z_rad - zz_slp1
            
            def objective(x, a, b, c, d):
                return a*x + b*x**2 + c*x**3 + d 
            popt, _ = curve_fit(objective, x, y)
            a, b, c, d = popt

            fig_rad = plt.figure(figsize=(2.5, 2.5))
            ax_rad = plt.axes(projection="3d")

            zz_rad = a*x + b*x**2 + c*x**3 + d
            Z_rad = griddata((x_rad, y_rad), zz_rad, (X_slp1, Y_slp1), method="cubic")

            my_fitting = np.polyfit(x, y, 3, full=True)
            coeff = my_fitting[0]
            SSE = my_fitting[1][0]
            diff = y - y.mean()
            square_diff = diff ** 2
            SST = square_diff.sum()
            R_sq2 = 1 - SSE/SST 

            btn_limitz_resi= btn_limitz3d
            top_limitz_resi = top_limitz3d

            Z_rad2 = Z_rad + np.mean(z3d) 
            ax_rad.contour3D(X_slp1, Y_slp1, Z_rad2, 300, cmap='turbo', alpha=1, antialiased=False)
            cset = ax_rad.contour(X_slp1, Y_slp1, Z_rad, 7, zdir='z', offset=btn_limitz_resi, linewidths=1,
                            cmap=cm.turbo, alpha=0.9)

            ax_rad.set_zlim(btn_limitz_resi, top_limitz_resi)                                                                                                                      
            ax_rad.view_init(30, 240)      
            ax_rad.set_xlabel('X (mm)')
            ax_rad.set_ylabel('Y (mm)')
            ax_rad.set_zlabel(unit)
            
            ax_rad.set_title(
                ' Abstracted radial pattern: \n z = (%.2f) r + (%.4f) r$^2$ + (%.6f) r$^3$ + (%.2f) \n R$^2$ = %.2f'
                % (a, b, c, d, R_sq2), loc='left', fontsize=5.5)
                 
            fig_rad.tight_layout(pad=1)    
            plot_rad ="radius"
            plt.savefig(plot_rad, bbox_inches='tight')
            img_rad = Image.open("radius.png")
            plot_rad_resized = img_rad.resize((356, 386))
            plot_rad_resized.save("radius_resized.png")     
            img_rad = ImageTk.PhotoImage(Image.open("radius_resized.png"))  
            self.canvas_graph.create_image(1180, 380, anchor="nw", image=img_rad)
            worksheet.insert_image('V16', 'radius.png', {'x_scale': 1.05, 'y_scale': 1.05})
            os.remove("radius_resized.png")
            plt.clf()
            plt.close(fig_rad)

# decompose_radius (2d) #4
            fig_rad2d = plt.figure(figsize=(2.22, 2.22))
            spec_rad2d = gridspec.GridSpec(ncols=1, nrows=1)
            ax_rad2d = fig_rad2d.add_subplot(spec_rad2d[0, 0])
            
            X_rad2d = X_slp1
            Y_rad2d = Y_slp1
            Z_rad2d = Z_rad2

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_bt_pre = self.limits_entry.get()
                limits_b = float(limits_bt_pre)
                limits_t_pre = self.limits_up.get()
                limits_t = float(limits_t_pre)
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_rad2d, Y_rad2d, Z_rad2d, levels = levels, cmap=plt.cm.turbo, alpha=0.95)
            else:
                limits_b = Min*0.97
                limits_t = Max*1.03
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_rad2d, Y_rad2d, Z_rad2d, levels = levels, cmap=plt.cm.turbo, alpha=0.95)

            x_rad2d = x_slp1 
            y_rad2d = y_slp1
            z_rad2d = zz_rad
            range_rad2d = round((max(z_rad2d) - min(z_rad2d)), 1)
            nonu_rad2d = round(0.5*100*(max(z_rad2d) - min(z_rad2d))/abs(Ave), 2)

            ax_rad2d.set_aspect('equal', adjustable='box')
            ax_rad2d.set_title(' \n Abstracted radial pattern: \n Range = %.1f unit, NonU = %.2f pct'
                            %(range_rad2d, nonu_rad2d), loc='left', fontsize=6)
            
            var3 = self.variable3.get()
            if (var3 == "Value"):
                for x_rad2d, y_rad2d, z_rad2d in zip(x_rad2d, y_rad2d, z_rad2d):
                    if value1 == '(A-B)/t' or value1 == '(B-A)/t' or value1 == 'A/t' or value1 == 'B/t':
                        label = "{:.0f}".format(z_rad2d*(60/time))
                    else:
                        label = "{:.0f}".format(z_rad2d)
                    plt.annotate(label, (x_rad2d, y_rad2d), textcoords="offset points", xytext=(0, -3), ha='center',
                            fontsize=4.3, alpha=0.85)
                    ax_rad2d.scatter(x_rad2d, y_rad2d, marker='o', s=1, color='k', alpha=0.0001)
            elif (var3 == "Dot"):
                for x_rad2d, y_rad2d, z_rad2d in zip(x_rad2d, y_rad2d, z_rad2d):
                    ax_rad2d.scatter(x_rad2d, y_rad2d, marker='o', s=1, color='k', alpha=0.3)
            elif (var3 == "Blank"):
                for x_rad2d, y_rad2d, z_rad2d in zip(x_rad2d, y_rad2d, z_rad2d):
                    ax_rad2d.scatter(x_rad2d, y_rad2d, marker='o', s=1, color='k', alpha=0.0001)
            else:
                for x_rad2d, y_rad2d, z_rad2d in zip(x_rad2d, y_rad2d, z_rad2d):
                    label = "{:.0f}".format(z_rad2d)
                    if (z_rad2d > 0):
                        ax_rad2d.scatter(x_rad2d, y_rad2d, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                    else:
                        ax_rad2d.scatter(x_rad2d, y_rad2d, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)

            ax_rad2d.set_xlabel('X (mm)')
            ax_rad2d.set_ylabel('Y (mm)')
            ax_rad2d.tick_params(axis='both', length=3)
            fig_rad2d.tight_layout(pad=0.1)
            plot_rad2d = "contour_rad2d"
            plt.savefig(plot_rad2d, bbox_inches='tight')        
            img_rad2d = Image.open("contour_rad2d.png")
            rad2d_resized = img_rad2d.resize((275, 297))
            rad2d_resized.save("rad2d_resized.png")     
            img_rad2d = ImageTk.PhotoImage(Image.open("rad2d_resized.png"))
            self.canvas_graph.create_image(1224, 1110, anchor="nw", image=img_rad2d)
            worksheet.insert_image('V43', 'contour_rad2d.png', {'x_scale': 0.95, 'y_scale': 0.97})
            os.remove("rad2d_resized.png")
            plt.clf()
            plt.close(fig_rad2d)

# decompose_residual (3d) #2
            fig_res = plt.figure(figsize=(2.5, 2.5))
            ax_res = plt.axes(projection="3d")
            
            degree_theta = float(self.rotation_entry.get())
            radian_rotation = math.radians(degree_theta)
            if (degree_theta == 0):
                x_res = df["A"]
                y_res = df["B"]
            else:
                x_orire = df["A"]
                y_orire = df["B"]
                x_res = math.cos(radian_rotation)*x_orire + math.sin(radian_rotation)*y_orire
                y_res = -math.sin(radian_rotation)*x_orire + math.cos(radian_rotation)*y_orire

            x_tmp = df[['x_tc', 'y_tc']].values.reshape(-1, 2)
            max_rows = max(len(x_slp1), len(y_slp1))
            x_ones = np.ones([max_rows, 1])
            x = np.append(x_ones, x_tmp, axis=1)
            y_tmp = z_slp1.tolist()
            y = np.array(y_tmp)

            def mse(coef, x, y):
                return np.mean((np.dot(x, coef) - y)**2)/2
            
            def gradients(coef, x, y):
                return np.mean(x.transpose()*(np.dot(x, coef) - y), axis=1)
            
            def multilinear_regression(coef, x, y, lr, b1=0.9, b2=0.999, epsilon=1e-8):
                prev_error = 0
                m_coef = np.zeros(coef.shape)
                v_coef = np.zeros(coef.shape)
                moment_m_coef = np.zeros(coef.shape)
                moment_v_coef = np.zeros(coef.shape)
                t = 0
                while True:
                    error = mse(coef, x, y)
                    if abs(error - prev_error) <= epsilon:
                        break
                    prev_error = error
                    grad = gradients(coef, x, y)
                    t += 1
                    m_coef = b1*m_coef + (1-b1)*grad
                    v_coef = b2*v_coef + (1-b2)*grad**2
                    moment_m_coef = m_coef / (1-b1**t)
                    moment_v_coef = v_coef / (1-b2**t)
                    delta = ((lr / moment_v_coef**0.5 + 1e-8)*(b1*moment_m_coef + (1-b1)*grad/(1-b1**t)))
                    coef = np.subtract(coef, delta)
                return coef
            
            coef = np.array([0, 0, 0])
            c = multilinear_regression(coef, x, y, 1e-1)
            
            x_res_grid = np.linspace(np.min(x_res), np.max(x_res), 120)
            y_res_grid = np.linspace(np.min(y_res), np.max(y_res), 120)
            X_res, Y_res = np.meshgrid(x_res_grid, y_res_grid)
            
            z_res = df["Z"]
            zz_res = c[1]*x_res + c[2]*y_res + c[0]
            Z_res = griddata((x_res, y_res), z_res, (X_res, Y_res), method="cubic")
            ZZ_res = griddata((x_res, y_res), zz_res, (X_res, Y_res), method="linear")

            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_z_pre = self.limits_entry.get()
                btn_limitz_res = float(limits_z_pre)
                limits_t_pre = self.limits_up.get()
                top_limitz_res = float(limits_t_pre)
            else:
                top_limitz_res = np.mean(z_res) + (max(z_res)-min(z_res))*1.5
                btn_limitz_res = np.mean(z_res) - (max(z_res)-min(z_res))*4.5
         
            ax_res.contour3D(X_res, Y_res, Z_res - ZZ_res + c[0] - Z_rad, 300, cmap='turbo', alpha=1,
                            antialiased=False)
            cset = ax_res.contour(X_res, Y_res, Z_res - ZZ_res + c[0] - Z_rad, 7, zdir='z', offset=btn_limitz_res,
                            linewidths=1, cmap=cm.turbo, alpha=1)

            z_after4 = z_res - zz_res + c[0] - zz_rad
            ave_res = round(np.mean(z_after4), 1)
            max_res = round(max(z_after4))
            min_res = round(min(z_after4))
            nonu_res = round(0.5*100*(max_res - min_res)/abs(ave_res), 2)
        
            ax_res.view_init(30, 240)
            ax_res.set_zlim(btn_limitz_res, top_limitz_res)
            ax_res.set_xlabel('X (mm)')
            ax_res.set_ylabel('Y (mm)')
            ax_res.set_zlabel(unit)
                 
            ax_res.set_title(
                            ' After tilting and radial pattern are corrected\n Mean = %.1f unit,  NonU = %.2f pct'
                            %(ave_res, nonu_res), loc='left', fontsize=6)
            fig_res.tight_layout(pad=1)    
            plot_res ="residual"
            plt.savefig(plot_res, bbox_inches='tight')
            img_res = Image.open("residual.png")
            plot_res_resized = img_res.resize((346, 356))
            plot_res_resized.save("residual_resized.png")     
            img_res = ImageTk.PhotoImage(Image.open("residual_resized.png"))
            self.canvas_graph.create_image(1180, 10, anchor="nw", image=img_res)
            worksheet.insert_image('V1', 'residual.png', {'x_scale': 1, 'y_scale': 1})
            os.remove("residual_resized.png")
            plt.clf()
            plt.close(fig_res)

# decompose_residual (2d) #2
            fig_res2 = plt.figure(figsize=(2.0, 2.0))
            spec_res2 = gridspec.GridSpec(ncols=1, nrows=1)
            ax_res2 = fig_res2.add_subplot(spec_res2[0, 0])

            X_res2 = X_res
            Y_res2 = Y_res
            Z_res2 = Z_res
            ZZ_res2 = ZZ_res
            Z7 = Z_res2 - ZZ_res2 - Z_rad + c[0]
            
            check_box = self.var_limits.get() 
            if (check_box == 1):
                limits_bt_pre = self.limits_entry.get()
                limits_b = float(limits_bt_pre)
                limits_t_pre = self.limits_up.get()
                limits_t = float(limits_t_pre)
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_res2, Y_res2, Z7, levels = levels, cmap=plt.cm.turbo, alpha=0.95)
            else:
                limits_b = Min*0.97
                limits_t = Max*1.03
                num = int(self.entry_contour.get())
                if (num >150):
                    num = 150
                    levels = np.linspace(limits_b, limits_t, num)
                else:
                    levels = np.linspace(limits_b, limits_t, num)
                cp = plt.contourf(X_res2, Y_res2, Z7, levels = levels, cmap=plt.cm.turbo, alpha=0.95)

            ax_res2.set_aspect('equal', adjustable='box')
            ax_res2.set_title(' \n After tilting and radial pattern corrected: \n Mean = %.1f unit,  NonU = %.2f pct'
                            %(ave_res, nonu_res), loc='left', fontsize=5.5)

            x_res2 = x_res
            y_res2 = y_res
            z_res2 = z_res
            zz_res2 = zz_res
            z7 = z_res2 - zz_res2 - zz_rad + c[0]

            var3 = self.variable3.get()
            if (var3 == "Value"):
                for x_res2, y_res2, z7 in zip(x_res2, y_res2, z7):
                    if value1 == '(A-B)/t' or value1 == '(B-A)/t' or value1 == 'A/t' or value1 == 'B/t':
                        label = "{:.0f}".format(z7*(60/time))
                    else:
                        label = "{:.0f}".format(z7)
                    plt.annotate(label, (x_res2, y_res2), textcoords="offset points", xytext=(0, -3), ha='center',
                            fontsize=4.3, alpha=0.85)
                    ax_res2.scatter(x_res2, y_res2, marker='o', s=1, color='k', alpha=0.0001)
            elif (var3 == "Dot"):
                for x_res2, y_res2, z7 in zip(x_res2, y_res2, z7):
                    ax_res2.scatter(x_res2, y_res2, marker='o', s=1, color='k', alpha=0.3)
            elif (var3 == "Blank"):
                for x_res2, y_res2, z7 in zip(x_res2, y_res2, z7):
                    ax_res2.scatter(x_res2, y_res2, marker='o', s=1, color='k', alpha=0.0001)
            else:
                for x_res2, y_res2, z7 in zip(x_res2, y_res2, z7):
                    label = "{:.0f}".format(z7)
                    if (z7 > Ave):
                        ax_res2.scatter(x_res2, y_res2, marker='+', s=13, linewidths=0.3, color='k', alpha=0.7)
                    else:
                        ax_res2.scatter(x_res2, y_res2, marker='_', s=13, linewidths=0.3, color='k', alpha=0.7)
            
            ax_res2.set_xlabel('X (mm)')
            ax_res2.set_ylabel('Y (mm)')
            ax_res2.tick_params(axis='both', length=3)
            fig_res2.tight_layout(pad=0.1)
            plot_res2 = "contour_res2"
            plt.savefig(plot_res2, bbox_inches='tight')        
            img_res2 = Image.open("contour_res2.png")
            res2_resized = img_res2.resize((288, 312))
            res2_resized.save("res2_resized.png")     
            img_res2 = ImageTk.PhotoImage(Image.open("res2_resized.png"))
            self.canvas_graph.create_image(1214, 780, anchor="nw", image=img_res2)
            worksheet.insert_image('V31', 'contour_res2.png', {'x_scale': 1.08, 'y_scale': 1.12})
            os.remove("res2_resized.png")
            plt.clf()
            plt.close(fig_res2)
            
        else:
            pass
        
        workbook.close()
        
        if (check_box2 == 1):
            os.remove("residual.png")
            os.remove("radius.png")
            os.remove("slope2.png")
            os.remove("slope1.png")
            os.remove("cdf.png")
            os.remove("hist.png")
            os.remove("scatter2.png")
            os.remove("scatter1.png")
            os.remove("3d.png")
            os.remove("cross_section2.png")
            os.remove("cross_section1.png")
            os.remove("contour.png")
            os.remove("contour_slp2d.png")
            os.remove("contour_skew.png")
            os.remove("contour_rad2d.png")
            os.remove("contour_res2.png")
        else:
            os.remove("3d.png")
            os.remove("cross_section2.png")
            os.remove("cross_section1.png")
            os.remove("contour.png")

        tk.mainloop()
                
def main():
    my_gui = app_gui()
main()       

# li.hou2009@gmail.com
# in memory of dear father Chongjian Hou (1928-2021) 
# https://houli1959.github.io/plasma/
